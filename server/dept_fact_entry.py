# -*- coding: utf-8 -*-
"""부서 주관 실적 입력: 평가배치(dept) 기준 목록 · 저장 · 엑셀."""
from __future__ import annotations

import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from fact_pipeline import INDICATOR_CODE_RE, eval_ym as to_eval_ym, refresh_facts, resolve_plan_set
from import_fact_upload import _actuals_equal, _write_change_log, normalize_eval_ym

SHEET_NAME = "부서실적"
HEADERS = ["평가월", "지표코드", "지표명", "실적", "주관부서", "단위"]
TEMPLATE_PATH = Path(__file__).resolve().parent / "dept_fact_entry_template.xlsx"


def list_distinct_depts(conn, year: int, month: int) -> list[str]:
    plan = resolve_plan_set(conn, year, month)
    if not plan:
        return []
    rows = conn.execute(
        """
        SELECT DISTINCT TRIM(dept) AS dept
        FROM eval_plan_item
        WHERE plan_set_id=? AND COALESCE(use_yn,'Y')='Y' AND TRIM(COALESCE(dept,'')) <> ''
        ORDER BY dept
        """,
        (plan["id"],),
    ).fetchall()
    return [r["dept"] for r in rows if r["dept"]]


def list_all_fact_entries(
    conn,
    year: int,
    month: int,
) -> dict[str, Any]:
    """해당월 평가배치 전체 지표 + fact_collect (Admin 전체 조회용)."""
    from fact_period import get_period_status

    ym = to_eval_ym(year, month)
    plan = resolve_plan_set(conn, year, month)
    period = get_period_status(conn, year, month)
    confirmed = {
        g["group_code"] for g in period.get("groups") or [] if g.get("status") == "confirmed"
    }
    frozen = period.get("period_status") == "frozen"
    if not plan:
        return {
            "year": year,
            "month": month,
            "eval_ym": ym,
            "scope": "all",
            "plan_set_id": None,
            "items": [],
            "period_status": period.get("period_status") or "open",
            "editable_for_non_admin": False,
        }

    rows = conn.execute(
        """
        SELECT
          e.indicator_code,
          MAX(e.label) AS label,
          MAX(e.dept) AS plan_dept,
          MAX(COALESCE(NULLIF(TRIM(ic.owner_group_code), ''), NULLIF(TRIM(cm.owner_group_code), ''))) AS ownership_group_code,
          MAX(COALESCE(NULLIF(TRIM(ic.dept), ''), NULLIF(TRIM(cm.dept), ''), e.dept, '')) AS ownership_dept,
          MAX(COALESCE(NULLIF(TRIM(e.unit), ''), NULLIF(TRIM(cm.unit), ''), ic.unit, '')) AS unit,
          MAX(e.collect_type) AS collect_type,
          MAX(e.mgmt_tool) AS mgmt_tool,
          GROUP_CONCAT(DISTINCT e.group_code) AS group_codes,
          GROUP_CONCAT(DISTINCT og.name) AS group_names,
          MAX(fc.actual) AS actual,
          MAX(fc.fetched_at) AS fetched_at,
          MAX(fc.batch_id) AS batch_id,
          CASE WHEN MAX(fc.id) IS NULL THEN 0 ELSE 1 END AS has_fact,
          MIN(e.sort_order) AS sort_order,
          MIN(e.group_code) AS sort_group
        FROM eval_plan_item e
        JOIN indicator_code ic ON ic.indicator_code = e.indicator_code
        JOIN indicator_common cm ON cm.common_code = ic.common_code
        JOIN owner_group og ON og.code = e.group_code
        LEFT JOIN fact_collect fc
          ON fc.indicator_code = e.indicator_code AND fc.eval_ym = ?
        WHERE e.plan_set_id = ?
          AND COALESCE(e.use_yn, 'Y') = 'Y'
        GROUP BY e.indicator_code
        ORDER BY sort_group, sort_order, e.indicator_code
        """,
        (ym, plan["id"]),
    ).fetchall()

    items = []
    for r in rows:
        gcodes = [x for x in str(r["group_codes"] or "").split(",") if x]
        group_locked = any(g in confirmed for g in gcodes)
        own_dept = r["ownership_dept"] or r["plan_dept"] or ""
        own_group = (r["ownership_group_code"] or "").strip().upper()
        items.append({
            "indicator_code": r["indicator_code"],
            "label": r["label"] or "",
            "unit": r["unit"] or "",
            "dept": own_dept,
            "ownership_dept": own_dept,
            "ownership_group_code": own_group,
            "collect_type": r["collect_type"] or "",
            "mgmt_tool": r["mgmt_tool"] or "",
            "group_codes": gcodes,
            "group_names": [x for x in str(r["group_names"] or "").split(",") if x],
            "actual": r["actual"],
            "fetched_at": r["fetched_at"],
            "batch_id": r["batch_id"],
            "has_fact": bool(r["has_fact"]),
            "manager_name": "",
            "group_confirmed": group_locked,
            "locked_for_non_admin": frozen or group_locked,
        })
    return {
        "year": year,
        "month": month,
        "eval_ym": ym,
        "scope": "all",
        "plan_set_id": plan["id"],
        "items": items,
        "period_status": period.get("period_status") or "open",
        "editable_for_non_admin": False,
    }


def list_dept_fact_entries(
    conn,
    year: int,
    month: int,
    *,
    dept: str,
) -> dict[str, Any]:
    """주관부서 = dept 인 평가배치 지표 + fact_collect (없으면 actual null)."""
    from fact_period import get_period_status

    dept = str(dept or "").strip()
    if not dept:
        raise ValueError("dept(주관부서)가 필요합니다")
    ym = to_eval_ym(year, month)
    plan = resolve_plan_set(conn, year, month)
    period = get_period_status(conn, year, month)
    confirmed = {
        g["group_code"] for g in period.get("groups") or [] if g.get("status") == "confirmed"
    }
    frozen = period.get("period_status") == "frozen"
    if not plan:
        return {
            "year": year,
            "month": month,
            "eval_ym": ym,
            "dept": dept,
            "plan_set_id": None,
            "items": [],
            "period_status": period.get("period_status") or "open",
            "editable_for_non_admin": not frozen,
        }

    rows = conn.execute(
        """
        SELECT
          e.indicator_code,
          MAX(e.label) AS label,
          MAX(e.dept) AS plan_dept,
          MAX(COALESCE(NULLIF(TRIM(ic.owner_group_code), ''), NULLIF(TRIM(cm.owner_group_code), ''))) AS ownership_group_code,
          MAX(COALESCE(NULLIF(TRIM(ic.dept), ''), NULLIF(TRIM(cm.dept), ''), e.dept, '')) AS ownership_dept,
          MAX(COALESCE(NULLIF(TRIM(e.unit), ''), NULLIF(TRIM(cm.unit), ''), ic.unit, '')) AS unit,
          MAX(e.collect_type) AS collect_type,
          MAX(e.mgmt_tool) AS mgmt_tool,
          GROUP_CONCAT(DISTINCT e.group_code) AS group_codes,
          GROUP_CONCAT(DISTINCT og.name) AS group_names,
          MAX(fc.actual) AS actual,
          MAX(fc.fetched_at) AS fetched_at,
          MAX(fc.batch_id) AS batch_id,
          CASE WHEN MAX(fc.id) IS NULL THEN 0 ELSE 1 END AS has_fact
        FROM eval_plan_item e
        JOIN indicator_code ic ON ic.indicator_code = e.indicator_code
        JOIN indicator_common cm ON cm.common_code = ic.common_code
        JOIN owner_group og ON og.code = e.group_code
        LEFT JOIN fact_collect fc
          ON fc.indicator_code = e.indicator_code AND fc.eval_ym = ?
        WHERE e.plan_set_id = ?
          AND COALESCE(e.use_yn, 'Y') = 'Y'
          AND TRIM(COALESCE(e.dept, '')) = ?
        GROUP BY e.indicator_code
        ORDER BY MAX(e.sort_order), e.indicator_code
        """,
        (ym, plan["id"], dept),
    ).fetchall()

    items = []
    for r in rows:
        gcodes = [x for x in str(r["group_codes"] or "").split(",") if x]
        group_locked = any(g in confirmed for g in gcodes)
        own_dept = r["ownership_dept"] or r["plan_dept"] or dept
        own_group = (r["ownership_group_code"] or "").strip().upper()
        items.append({
            "indicator_code": r["indicator_code"],
            "label": r["label"] or "",
            "unit": r["unit"] or "",
            "dept": own_dept,
            "ownership_dept": own_dept,
            "ownership_group_code": own_group,
            "collect_type": r["collect_type"] or "",
            "mgmt_tool": r["mgmt_tool"] or "",
            "group_codes": gcodes,
            "group_names": [x for x in str(r["group_names"] or "").split(",") if x],
            "actual": r["actual"],
            "fetched_at": r["fetched_at"],
            "batch_id": r["batch_id"],
            "has_fact": bool(r["has_fact"]),
            "manager_name": "",  # 마스터에 담당자 필드 없음 — UI에서 입력자 표시
            "group_confirmed": group_locked,
            "locked_for_non_admin": frozen or group_locked,
        })
    return {
        "year": year,
        "month": month,
        "eval_ym": ym,
        "dept": dept,
        "plan_set_id": plan["id"],
        "items": items,
        "period_status": period.get("period_status") or "open",
        "editable_for_non_admin": (not frozen) and any(
            not it["locked_for_non_admin"] for it in items
        ) if items else (not frozen),
    }


def list_group_fact_entries(
    conn,
    year: int,
    month: int,
    *,
    group_code: str,
) -> dict[str, Any]:
    """평가배치 그룹 소관 지표 + fact_collect."""
    from fact_period import get_period_status

    group_code = str(group_code or "").strip().upper()
    if not group_code:
        raise ValueError("group_code가 필요합니다")
    ym = to_eval_ym(year, month)
    plan = resolve_plan_set(conn, year, month)
    period = get_period_status(conn, year, month)
    gstat = next((g for g in period["groups"] if g["group_code"] == group_code), None)
    if not plan:
        return {
            "year": year, "month": month, "eval_ym": ym, "group_code": group_code,
            "plan_set_id": None, "items": [],
            "period_status": period["period_status"],
            "group_confirm_status": (gstat or {}).get("status") or "open",
            "editable": period["period_status"] != "frozen",
        }
    rows = conn.execute(
        """
        SELECT
          e.indicator_code,
          MAX(e.label) AS label,
          MAX(e.dept) AS plan_dept,
          MAX(COALESCE(NULLIF(TRIM(ic.owner_group_code), ''), NULLIF(TRIM(cm.owner_group_code), ''))) AS ownership_group_code,
          MAX(COALESCE(NULLIF(TRIM(ic.dept), ''), NULLIF(TRIM(cm.dept), ''), e.dept, '')) AS ownership_dept,
          MAX(COALESCE(NULLIF(TRIM(e.unit), ''), NULLIF(TRIM(cm.unit), ''), ic.unit, '')) AS unit,
          MAX(e.collect_type) AS collect_type,
          MAX(e.mgmt_tool) AS mgmt_tool,
          MAX(e.group_code) AS group_code,
          MAX(og.name) AS group_name,
          MAX(fc.actual) AS actual,
          MAX(fc.fetched_at) AS fetched_at,
          MAX(fc.batch_id) AS batch_id,
          CASE WHEN MAX(fc.id) IS NULL THEN 0 ELSE 1 END AS has_fact
        FROM eval_plan_item e
        JOIN indicator_code ic ON ic.indicator_code = e.indicator_code
        JOIN indicator_common cm ON cm.common_code = ic.common_code
        JOIN owner_group og ON og.code = e.group_code
        LEFT JOIN fact_collect fc
          ON fc.indicator_code = e.indicator_code AND fc.eval_ym = ?
        WHERE e.plan_set_id = ?
          AND COALESCE(e.use_yn, 'Y') = 'Y'
          AND e.group_code = ?
        GROUP BY e.indicator_code
        ORDER BY MAX(e.sort_order), e.indicator_code
        """,
        (ym, plan["id"], group_code),
    ).fetchall()
    items = []
    for r in rows:
        own_dept = r["ownership_dept"] or r["plan_dept"] or ""
        own_group = (r["ownership_group_code"] or "").strip().upper()
        items.append({
            "indicator_code": r["indicator_code"],
            "label": r["label"] or "",
            "unit": r["unit"] or "",
            "dept": own_dept,
            "ownership_dept": own_dept,
            "ownership_group_code": own_group,
            "collect_type": r["collect_type"] or "",
            "mgmt_tool": r["mgmt_tool"] or "",
            "group_codes": [group_code],
            "group_names": [r["group_name"] or group_code],
            "actual": r["actual"],
            "fetched_at": r["fetched_at"],
            "batch_id": r["batch_id"],
            "has_fact": bool(r["has_fact"]),
            "manager_name": "",
        })
    g_confirmed = (gstat or {}).get("status") == "confirmed"
    frozen = period["period_status"] == "frozen"
    for it in items:
        it["group_confirmed"] = g_confirmed
        it["locked_for_non_admin"] = frozen or g_confirmed
    return {
        "year": year,
        "month": month,
        "eval_ym": ym,
        "group_code": group_code,
        "plan_set_id": plan["id"],
        "items": items,
        "period_status": period["period_status"],
        "group_confirm_status": (gstat or {}).get("status") or "open",
        "editable_for_non_admin": (not frozen) and (not g_confirmed),
    }


def _assert_codes_belong_to_dept(conn, year: int, month: int, dept: str, codes: list[str]) -> None:
    catalog = list_dept_fact_entries(conn, year, month, dept=dept)
    allowed = {it["indicator_code"] for it in catalog["items"]}
    bad = [c for c in codes if c not in allowed]
    if bad:
        raise ValueError(f"주관부서({dept}) 소관이 아닌 지표코드: {', '.join(bad[:10])}")


def _assert_codes_belong_to_group(conn, year: int, month: int, group_code: str, codes: list[str]) -> None:
    catalog = list_group_fact_entries(conn, year, month, group_code=group_code)
    allowed = {it["indicator_code"] for it in catalog["items"]}
    bad = [c for c in codes if c not in allowed]
    if bad:
        raise ValueError(f"그룹({group_code}) 소관이 아닌 지표코드: {', '.join(bad[:10])}")


def _assert_codes_belong_to_plan(conn, year: int, month: int, codes: list[str]) -> dict[str, list[str]]:
    """전체 스코프: 평가배치에 있는 코드인지 확인하고 코드→그룹 맵 반환."""
    catalog = list_all_fact_entries(conn, year, month)
    by_code = {it["indicator_code"]: it for it in catalog["items"]}
    bad = [c for c in codes if c not in by_code]
    if bad:
        raise ValueError(f"평가배치에 없는 지표코드: {', '.join(bad[:10])}")
    return {c: (by_code[c].get("group_codes") or []) for c in codes}


def save_dept_fact_entries(
    conn,
    year: int,
    month: int,
    *,
    dept: str = "",
    group_code: str = "",
    scope_all: bool = False,
    updates: list[dict[str, Any]],
    acted_by: str = "ui",
    actor_role: str = "",
    refresh: bool = True,
) -> dict[str, Any]:
    """부서/그룹/전체 스코프 실적 저장 → fact_collect + 변경로그 + 산출 재계산."""
    from fact_period import assert_writable

    dept = str(dept or "").strip()
    group_code = str(group_code or "").strip().upper()
    scope_all = bool(scope_all)
    if not scope_all and not dept and not group_code:
        raise ValueError("dept 또는 group_code가 필요합니다")
    if scope_all and str(actor_role or "").strip().lower() != "admin":
        raise ValueError("전체 실적 저장은 Admin만 가능합니다")
    ym = to_eval_ym(year, month)
    started = datetime.now(timezone.utc).isoformat()

    normalized: list[dict] = []
    for u in updates or []:
        code = str(u.get("indicator_code") or u.get("indicatorCode") or "").strip().upper()
        if not code:
            continue
        if not INDICATOR_CODE_RE.match(code):
            raise ValueError(f"지표코드 형식 오류: {code}")
        raw = u.get("actual") if "actual" in u else u.get("value")
        if raw is None or str(raw).strip() == "":
            raise ValueError(f"{code}: 실적 값이 필요합니다")
        actual = float(raw)
        normalized.append({"indicator_code": code, "actual": actual})

    if not normalized:
        raise ValueError("저장할 실적 행이 없습니다")

    code_groups: dict[str, list[str]] = {}
    if scope_all:
        code_groups = _assert_codes_belong_to_plan(
            conn, year, month, [x["indicator_code"] for x in normalized],
        )
        scope_groups: list[str] = []
        for gs in code_groups.values():
            scope_groups.extend(gs)
        scope_label = "all"
    elif group_code:
        _assert_codes_belong_to_group(conn, year, month, group_code, [x["indicator_code"] for x in normalized])
        scope_groups = [group_code]
        scope_label = f"group:{group_code}"
    else:
        _assert_codes_belong_to_dept(conn, year, month, dept, [x["indicator_code"] for x in normalized])
        catalog = list_dept_fact_entries(conn, year, month, dept=dept)
        by_code = {it["indicator_code"]: it for it in catalog["items"]}
        scope_groups = []
        for it in normalized:
            gs = by_code.get(it["indicator_code"], {}).get("group_codes") or []
            code_groups[it["indicator_code"]] = gs
            scope_groups.extend(gs)
            if not gs:
                scope_groups.append(it["indicator_code"].split("-")[-1])
        scope_label = f"dept:{dept}"

    assert_writable(conn, year, month, group_codes=scope_groups, actor_role=actor_role)

    # upload batch for audit trail
    cur = conn.execute(
        """
        INSERT INTO fact_upload_batch(
          filename, status, counts_json, error_text, uploaded_by, created_at
        ) VALUES (?,?,?,?,?,?)
        """,
        (
            f"entry:{scope_label}:{ym}",
            "running",
            "{}",
            "",
            acted_by,
            started,
        ),
    )
    batch_id = cur.lastrowid

    cur_sb = conn.execute(
        "INSERT INTO sync_batch(mode, eval_ym, status, started_at) VALUES (?,?,?,?)",
        ("dept_entry", ym, "running", started),
    )
    sync_id = cur_sb.lastrowid

    changed = 0
    new_rows = 0
    same = 0
    for it in normalized:
        existing = conn.execute(
            "SELECT actual FROM fact_collect WHERE eval_ym=? AND indicator_code=?",
            (ym, it["indicator_code"]),
        ).fetchone()
        prev = float(existing["actual"]) if existing and existing["actual"] is not None else None
        if existing is None:
            kind = "new"
            new_rows += 1
        elif _actuals_equal(prev, it["actual"]):
            kind = "same"
            same += 1
        else:
            kind = "changed"
            changed += 1

        gs = code_groups.get(it["indicator_code"]) or []
        group = group_code or (gs[0] if gs else (
            it["indicator_code"].split("-")[-1] if "-" in it["indicator_code"] else ""
        ))
        conn.execute(
            """
            INSERT INTO fact_upload_item(
              batch_id, eval_ym, indicator_code, group_code, actual, prev_actual,
              change_kind, status, error_text, export_status, row_no
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                batch_id, ym, it["indicator_code"], group, it["actual"], prev,
                kind, "ok", "", "pending", None,
            ),
        )
        _write_change_log(
            conn,
            batch_id=batch_id,
            eval_ym=ym,
            indicator_code=it["indicator_code"],
            group_code=group,
            prev_actual=prev,
            new_actual=it["actual"],
            change_kind=kind,
            action="dept_entry",
            acted_by=acted_by,
        )
        conn.execute(
            """
            INSERT INTO fact_collect(eval_ym, indicator_code, actual, batch_id, fetched_at)
            VALUES (?,?,?,?,?)
            ON CONFLICT(eval_ym, indicator_code) DO UPDATE SET
              actual=excluded.actual,
              batch_id=excluded.batch_id,
              fetched_at=excluded.fetched_at
            """,
            (ym, it["indicator_code"], it["actual"], sync_id, started),
        )

    conn.execute(
        """
        UPDATE sync_batch SET status=?, counts_json=?, finished_at=? WHERE id=?
        """,
        (
            "ok",
            json.dumps({"source": "dept_entry", "scope": scope_label, "rows": len(normalized)}, ensure_ascii=False),
            datetime.now(timezone.utc).isoformat(),
            sync_id,
        ),
    )

    refreshed = []
    errors = []
    if refresh:
        try:
            refresh_facts(conn, year, month, skip_collect=True)
            refreshed.append(ym)
        except Exception as e:
            errors.append(f"refresh {ym} 실패: {e}")

    status = "ok" if not errors else "partial"
    counts = {
        "rows_ok": len(normalized),
        "rows_new": new_rows,
        "rows_changed": changed,
        "rows_same": same,
        "dept": dept,
        "group_code": group_code,
        "scope": scope_label,
        "yms": [ym],
        "refreshed": refreshed,
    }
    conn.execute(
        """
        UPDATE fact_upload_batch
        SET status=?, counts_json=?, error_text=?, finished_at=?
        WHERE id=?
        """,
        (
            status,
            json.dumps(counts, ensure_ascii=False),
            "\n".join(errors),
            datetime.now(timezone.utc).isoformat(),
            batch_id,
        ),
    )
    conn.commit()
    return {
        "ok": True,
        "batch_id": batch_id,
        "status": status,
        "counts": counts,
        "errors": errors,
        "eval_ym": ym,
        "dept": dept,
        "group_code": group_code,
        "scope": scope_label,
    }


def write_dept_entry_workbook(
    conn,
    year: int,
    month: int,
    *,
    dept: str = "",
    group_code: str = "",
    scope_all: bool = False,
    path: Path | None = None,
) -> Path:
    group_code = str(group_code or "").strip().upper()
    dept = str(dept or "").strip()
    if scope_all:
        data = list_all_fact_entries(conn, year, month)
        scope = "all"
    elif group_code:
        data = list_group_fact_entries(conn, year, month, group_code=group_code)
        scope = group_code
    else:
        data = list_dept_fact_entries(conn, year, month, dept=dept)
        scope = dept or "dept"
    ym = data["eval_ym"]
    safe = re.sub(r"[^\w\-]+", "_", scope)[:40] or "scope"
    target = Path(path) if path else (
        Path(__file__).resolve().parent / "data" / f"dept_fact_{safe}_{ym}.xlsx"
    )
    target.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(HEADERS)
    for it in data["items"]:
        ws.append([
            ym,
            it["indicator_code"],
            it["label"],
            it["actual"] if it["actual"] is not None else "",
            it["dept"],
            it["unit"],
        ])
    guide = wb.create_sheet("안내")
    guide.append(["컬럼", "설명"])
    guide.append(["평가월", "YYYYMM (업로드 시 선택 연월과 일치해야 함)"])
    guide.append(["지표코드", "필수"])
    guide.append(["지표명", "참고용 (업로드 시 무시)"])
    guide.append(["실적", "필수. 숫자"])
    guide.append(["주관부서", "참고용"])
    guide.append(["단위", "참고용"])
    wb.save(target)
    return target


def parse_dept_entry_workbook(path: Path, *, expect_ym: str | None = None) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("빈 엑셀입니다")
    header = [str(c or "").strip() for c in rows[0]]
    col: dict[str, int] = {}
    aliases = {
        "평가월": ("평가월", "eval_ym", "YM"),
        "지표코드": ("지표코드", "indicator_code", "코드"),
        "실적": ("실적", "actual", "값"),
    }
    for field, names in aliases.items():
        for i, h in enumerate(header):
            if h in names:
                col[field] = i
                break
    if "지표코드" not in col or "실적" not in col:
        raise ValueError("필수 컬럼 없음: 지표코드, 실적")

    by_key: dict[str, dict] = {}
    for ridx, row in enumerate(rows[1:], start=2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        code = str(row[col["지표코드"]] or "").strip().upper()
        if not code:
            raise ValueError(f"행 {ridx}: 지표코드 없음")
        if not INDICATOR_CODE_RE.match(code):
            raise ValueError(f"행 {ridx}: 지표코드 형식 오류 {code}")
        if "평가월" in col:
            ym = normalize_eval_ym(row[col["평가월"]])
            if expect_ym and ym != expect_ym:
                raise ValueError(f"행 {ridx}: 평가월 {ym} ≠ 선택월 {expect_ym}")
        raw = row[col["실적"]]
        if raw is None or str(raw).strip() == "":
            raise ValueError(f"행 {ridx}: 실적 없음")
        by_key[code] = {"indicator_code": code, "actual": float(raw)}
    if not by_key:
        raise ValueError("유효한 행이 없습니다")
    return list(by_key.values())
