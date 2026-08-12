# -*- coding: utf-8 -*-
"""kpi.sqlite 전체 테이블을 엑셀로 백업."""
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from db import get_connection

ROOT = Path(__file__).resolve().parent
OUT_DIR = ROOT / "backups"


def _cell(v):
    if v is None:
        return ""
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False)
    if isinstance(v, bytes):
        return v.hex()
    return v


def _autosize(ws, max_width=48, sample=80):
    for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=min(ws.max_row, sample), max_col=ws.max_column), 1):
        width = 8
        for cell in col:
            if cell.value is None:
                continue
            width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def main() -> Path:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = OUT_DIR / f"kpi_demo_data_backup_{stamp}.xlsx"

    conn = get_connection()
    try:
        tables = [
            r[0]
            for r in conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY 1"
            ).fetchall()
        ]
        wb = Workbook()
        default = wb.active
        default.title = "_manifest"
        default.append(["table", "rows", "columns"])

        summary = []
        used_sheet_names: set[str] = set()
        for name in tables:
            cur = conn.execute(f'SELECT * FROM "{name}"')
            cols = [d[0] for d in cur.description]
            rows = cur.fetchall()
            summary.append((name, len(rows), len(cols)))

            # Excel sheet name max 31 chars; keep unique
            base = name[:31]
            sheet_name = base
            n = 2
            while sheet_name in used_sheet_names:
                suffix = f"_{n}"
                sheet_name = f"{base[: 31 - len(suffix)]}{suffix}"
                n += 1
            used_sheet_names.add(sheet_name)

            ws = wb.create_sheet(title=sheet_name)
            ws.append(cols)
            for row in rows:
                ws.append([_cell(row[c]) for c in cols])
            if rows:
                _autosize(ws)

        for name, n_rows, n_cols in summary:
            default.append([name, n_rows, n_cols])
        default.append([])
        default.append(["exported_at", datetime.now().isoformat(timespec="seconds")])
        default.append(["db_path", str(ROOT / "data" / "kpi.sqlite")])
        _autosize(default)

        wb.save(out_path)
        print(f"OK {out_path}")
        print(f"tables={len(tables)} total_rows={sum(r for _, r, _ in summary)}")
        for name, n_rows, n_cols in summary:
            print(f"  {name}: {n_rows} rows, {n_cols} cols")
        return out_path
    finally:
        conn.close()


if __name__ == "__main__":
    main()
