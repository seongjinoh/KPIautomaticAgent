from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

SHEET_NAME = "eval_plan_items"
TEMPLATE_PATH = Path(__file__).resolve().parent / "eval_plan_upload_template.xlsx"

# 평가배치 전용 체계 (코드마스터 분류와 별개)
HEADERS = [
    "indicator_code",
    "group_code",
    "순서",
    "평가Lv1",
    "평가Lv2",
    "평가Lv3",
    "표시명",
    "단위",
    "가중치",
    "Core",
    "산식구분",
    "연간목표",
    "기준실적",
    "목표방향",
    *[f"{m}월목표" for m in range(1, 13)],
    "상한",
    "하한",
    "기본승수",
    "조정승수",
    "비고",
    "custom_achievement_expr",
    *[f"Filter{i}" for i in range(1, 31)],
]

SAMPLE_ROWS = [
    [
        "CAP-0001-0001-RAT-CIG", "CIG", 1,
        "본원적 수익력", "조정 ROC", "조정 ROC", "조정 ROC",
        "%", 5, "Y", "flat", 0.2236, 0.2, "increase",
        *([""] * 12),
        130, 70, 1, 1, "",
        "",
        *([""] * 30),
    ],
    [
        "CUS-0002-0002-NET-SG1", "SG1", 2,
        "고객", "활성기반", "MAU", "MAU",
        "명", 3, "N", "custom", 12800, 9500, "increase",
        9800, 10050, 10300, 10500, 10800, 11000, 11200, 11500, 11800, 12000, 12300, 12800,
        "", "", 1, 1, "",
        "actual / target * 100",
        *([""] * 30),
    ],
    [
        "CAP-0001-0003-RAT-CIG", "CIG", 3,
        "본원적 수익력", "RORWA", "RORWA", "RORWA",
        "%", 3, "Y", "linear", 0.15, 0.12, "increase",
        *([""] * 12),
        120, 80, 1, 1, "",
        "",
        *([""] * 30),
    ],
]


def _core_yn_from_value(value, default: str = "N") -> str:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return "Y" if value else "N"
    s = str(value).strip().upper()
    if s in ("Y", "YES", "1", "TRUE", "T", "CORE", "O", "예"):
        return "Y"
    if s in ("N", "NO", "0", "FALSE", "F", "X", "아니오", "아님"):
        return "N"
    return default


def _core_yn_from_item(item: dict, default: str = "N") -> str:
    for k in ("is_core", "isCore", "Core", "core"):
        if k in item and item.get(k) is not None and item.get(k) != "":
            return _core_yn_from_value(item.get(k), default)
    return default


def write_template(path: Path = TEMPLATE_PATH) -> Path:
    return write_eval_workbook(SAMPLE_ROWS, path, include_guide=True)


def _parse_filters(raw) -> dict:
    if not raw:
        return {}
    if isinstance(raw, dict):
        return {str(k): v for k, v in raw.items() if v is not None and str(v).strip() != ""}
    if isinstance(raw, str):
        try:
            data = json.loads(raw)
            if isinstance(data, dict):
                return {str(k): v for k, v in data.items() if v is not None and str(v).strip() != ""}
        except Exception:
            return {}
    return {}


def _parse_monthly(raw) -> dict:
    if not raw:
        return {}
    if isinstance(raw, dict):
        return {str(k): v for k, v in raw.items()}
    if isinstance(raw, str):
        try:
            data = json.loads(raw)
            if isinstance(data, dict):
                return {str(k): v for k, v in data.items()}
        except Exception:
            return {}
    return {}


def item_to_row(item: dict) -> list:
    monthly = _parse_monthly(
        item.get("custom_monthly_targets_json")
        or item.get("customMonthlyTargets")
        or item.get("custom_monthly_targets")
    )
    filters = _parse_filters(
        item.get("filters_json")
        or item.get("filters")
        or item.get("filtersJson")
    )
    month_vals = []
    for m in range(1, 13):
        v = monthly.get(str(m), monthly.get(m, ""))
        month_vals.append("" if v is None or v == "" else v)

    filter_vals = []
    for i in range(1, 31):
        v = filters.get(str(i), filters.get(i, ""))
        filter_vals.append("" if v is None else v)

    return [
        str(item.get("indicator_code") or item.get("indicatorCode") or "").strip().upper(),
        str(item.get("group_code") or item.get("groupCode") or "").strip().upper(),
        item.get("sort_order") if item.get("sort_order") is not None else item.get("sortOrder") or "",
        item.get("eval_category_lv1") or item.get("evalCategoryLv1") or "",
        item.get("eval_category_lv2") or item.get("evalCategoryLv2") or "",
        item.get("eval_category_lv3") or item.get("evalCategoryLv3") or "",
        item.get("label") or "",
        item.get("unit") or "",
        item.get("weight") if item.get("weight") is not None else "",
        _core_yn_from_item(item),
        str(item.get("achievement_mode") or item.get("achievementMode") or "linear").strip().lower(),
        item.get("annual_target") if item.get("annual_target") is not None else item.get("annualTarget") or "",
        item.get("baseline_actual") if item.get("baseline_actual") is not None else item.get("baselineActual") or "",
        str(item.get("goal_direction") or item.get("goalDirection") or "increase").strip().lower(),
        *month_vals,
        item.get("cap_max") if item.get("cap_max") is not None else item.get("capMax") or "",
        item.get("cap_min") if item.get("cap_min") is not None else item.get("capMin") or "",
        item.get("score_rule") if item.get("score_rule") is not None else item.get("scoreRule") or "",
        item.get("penalty_rule") if item.get("penalty_rule") is not None else item.get("penaltyRule") or "",
        item.get("remark") or "",
        item.get("custom_achievement_expr") or item.get("customAchievementExpr") or "",
        *filter_vals,
    ]


def write_eval_workbook(rows: list, path: Path, include_guide: bool = True) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(HEADERS)
    for row in rows:
        if len(row) != len(HEADERS):
            raise ValueError(f"row length {len(row)} != headers {len(HEADERS)}")
        ws.append(row)

    if include_guide:
        guide = wb.create_sheet("guide")
        guide.append(["field", "description"])
        guide.append(["indicator_code", "필수. 코드체계 지표코드 (코드마스터)"])
        guide.append(["group_code", "필수. indicator_code 소속 그룹코드"])
        guide.append(["평가Lv1~Lv3", "평가배치 전용 하이어러키 (코드마스터 Lv와 별개)"])
        guide.append(["표시명", "해당 연도 평가 표시명(Label)"])
        guide.append(["가중치", "평가 비중(%)"])
        guide.append(["Core", "핵심지표 지정. Y/N (연도별 평가배치 시 수동 지정. 비중 상위와 무관)"])
        guide.append(["산식구분", "linear / flat / custom"])
        guide.append(["연간목표", "연간 목표. flat 은 매월 동일 비교, linear 는 일수 안분 기준"])
        guide.append(["기준실적", "linear 누적 월목표용 baseline (연초/기준)"])
        guide.append(["1~12월목표", "컬럼은 항상 존재. custom 일 때만 저장, linear/flat 은 무시"])
        guide.append(["상한/하한", "달성률 상·하한 (숫자)"])
        guide.append(["기본승수/조정승수", "숫자 승수"])
        guide.append(["Filter1~30", "custom 식에서 filter1..filter30 참조. 숫자 또는 지표코드"])
        guide.append(["custom_achievement_expr", "custom 달성률 식. 변수: actual, target, filter1..30"])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def export_eval_items(items: list[dict], path: Path) -> Path:
    rows = [item_to_row(item) for item in (items or [])]
    return write_eval_workbook(rows, path, include_guide=True)


def _rows(ws):
    it = ws.iter_rows(values_only=True)
    headers = [str(c).strip() if c is not None else "" for c in next(it)]
    rows = []
    for raw in it:
        if not any(v is not None and str(v).strip() != "" for v in raw):
            continue
        row = {headers[i]: raw[i] for i in range(len(headers)) if headers[i]}
        rows.append(row)
    return rows


def _first(row: dict, *keys):
    for k in keys:
        if k in row and row.get(k) is not None and str(row.get(k)).strip() != "":
            return row.get(k)
    return None


def _mode_from_cell(value) -> str:
    raw = str(value or "linear").strip().lower()
    if raw in ("flat", "custom", "linear"):
        return raw
    if "flat" in raw:
        return "flat"
    if "custom" in raw or "커스텀" in raw:
        return "custom"
    return "linear"


def parse_eval_workbook(xlsx_path: Path) -> list[dict]:
    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(str(path))
    wb = load_workbook(path, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"missing sheet: {SHEET_NAME}")

    items = []
    errors = []
    for idx, row in enumerate(_rows(wb[SHEET_NAME]), start=2):
        indicator_code = str(_first(row, "indicator_code", "코드") or "").strip().upper()
        group_code = str(_first(row, "group_code") or "").strip().upper()
        if not indicator_code:
            errors.append(f"row {idx}: indicator_code required")
            continue
        if not group_code:
            errors.append(f"row {idx}: group_code required")
            continue
        mode = _mode_from_cell(_first(row, "산식구분", "achievement_mode"))
        try:
            weight = float(_first(row, "가중치", "weight") or 0)
        except Exception:
            errors.append(f"row {idx}: weight must be numeric")
            continue

        monthly = {}
        if mode == "custom":
            for m in range(1, 13):
                v = row.get(f"{m}월목표")
                if v is not None and str(v).strip() != "":
                    try:
                        monthly[str(m)] = float(v)
                    except Exception:
                        errors.append(f"row {idx}: {m}월목표 must be numeric")

        filters = {}
        for i in range(1, 31):
            v = row.get(f"Filter{i}")
            if v is not None and str(v).strip() != "":
                filters[str(i)] = str(v).strip()

        direction = str(_first(row, "목표방향", "goal_direction") or "increase").strip().lower()
        if direction in ("decrease", "감소", "낮을수록"):
            direction = "decrease"
        else:
            direction = "increase"

        score = _first(row, "기본승수", "score_rule", "scoreRule", "배점기준")
        penalty = _first(row, "조정승수", "penalty_rule", "penaltyRule", "감점기준")

        items.append({
            "indicator_code": indicator_code,
            "group_code": group_code,
            "mgmt_tool": "KPI",
            "eval_category_lv1": str(_first(row, "평가Lv1", "eval_category_lv1", "evalCategoryLv1", "분야") or "").strip(),
            "eval_category_lv2": str(_first(row, "평가Lv2", "eval_category_lv2", "evalCategoryLv2", "세부분야") or "").strip(),
            "eval_category_lv3": str(_first(row, "평가Lv3", "eval_category_lv3", "evalCategoryLv3") or "").strip(),
            "label": str(_first(row, "표시명", "label", "지표명") or "").strip(),
            "weight": weight,
            "is_core": _core_yn_from_value(_first(row, "Core", "is_core", "isCore", "core")),
            "annual_target": float(_first(row, "연간목표", "annual_target", "annualTarget") or 0),
            "unit": str(_first(row, "단위", "unit") or "").strip(),
            "collect_type": "",
            "dept": "",
            "data_source": "",
            "h1_target": None,
            "h2_target": None,
            "score_rule": "" if score is None else str(score).strip(),
            "penalty_rule": "" if penalty is None else str(penalty).strip(),
            "cap_max": _first(row, "상한", "cap_max", "capMax"),
            "cap_min": _first(row, "하한", "cap_min", "capMin"),
            "remark": str(_first(row, "비고", "remark") or "").strip(),
            "achievement_mode": mode,
            "goal_direction": direction,
            "baseline_actual": float(_first(row, "기준실적", "baseline_actual", "baselineActual") or 0),
            "custom_achievement_expr": str(_first(row, "custom_achievement_expr", "customAchievementExpr") or "").strip(),
            "custom_monthly_targets_json": json.dumps(monthly, ensure_ascii=False) if monthly else None,
            "filters_json": json.dumps(filters, ensure_ascii=False) if filters else None,
            "sort_order": int(_first(row, "순서", "sort_order", "sortOrder") or idx - 1),
        })

    if errors:
        raise ValueError("; ".join(errors))
    return items


if __name__ == "__main__":
    out = write_template()
    print(f"wrote {out}")
