# -*- coding: utf-8 -*-
"""최종 엑셀 코드마스터 → SQLite 임포트 (01~04만)."""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook

from db import compose_indicator_code, counts, get_connection, init_schema
from org_group import normalize_org_level

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = ROOT / "은행_KPI_3단분류_코드마스터_최종.xlsx"

SHEET_GROUP = "01_그룹마스터"
SHEET_LV = "02_분류체계"
SHEET_COMMON = "03_지표마스터"
SHEET_CODES = "04_그룹별코드"


def _yn(value) -> str:
    s = str(value or "").strip()
    if s in ("사용", "Y", "y", "1", "공통"):
        return "Y"
    if s in ("미사용", "N", "n", "0"):
        return "N"
    return "Y" if s else "Y"


def _rows(ws):
    it = ws.iter_rows(values_only=True)
    hdr = [str(c).strip() if c is not None else "" for c in next(it)]
    data = []
    for r in it:
        if not any(v is not None and str(v).strip() != "" for v in r):
            continue
        data.append({hdr[i]: r[i] for i in range(len(hdr)) if hdr[i]})
    return data


def clear_code_tables(conn) -> None:
    conn.execute("DELETE FROM indicator_code")
    conn.execute("DELETE FROM indicator_common")
    conn.execute("DELETE FROM code_lv2")
    conn.execute("DELETE FROM code_lv1")
    conn.execute("DELETE FROM owner_group")


def import_workbook(xlsx_path: Path, conn=None) -> dict:
    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(str(path))

    own = conn is None
    if own:
        init_schema()
        conn = get_connection()

    wb = load_workbook(path, read_only=True, data_only=True)
    for name in (SHEET_GROUP, SHEET_LV, SHEET_COMMON, SHEET_CODES):
        if name not in wb.sheetnames:
            raise ValueError(f"missing sheet: {name}")

    groups = _rows(wb[SHEET_GROUP])
    lv_rows = _rows(wb[SHEET_LV])
    commons = _rows(wb[SHEET_COMMON])
    codes = _rows(wb[SHEET_CODES])

    errors = []
    clear_code_tables(conn)

    for r in groups:
        code = str(r.get("그룹코드") or "").strip().upper()
        name = str(r.get("그룹명") or "").strip()
        if not code:
            continue
        sort_order = int(r.get("정렬순서") or 0)
        org_level = normalize_org_level(r.get("조직레벨") or r.get("org_level") or "GROUP")
        parent_code = str(r.get("상위코드") or r.get("parent_code") or "").strip().upper() or None
        if org_level == "BANK":
            parent_code = None
        conn.execute(
            """INSERT INTO owner_group(code, name, sort_order, use_yn, org_level, parent_code)
               VALUES (?,?,?,?,?,?)""",
            (code, name, sort_order, _yn(r.get("사용여부")), org_level, parent_code),
        )

    lv1_seen = {}
    lv2_seen = {}
    for r in lv_rows:
        lv1 = str(r.get("Lv1코드") or "").strip().upper()
        lv1_name = str(r.get("Lv1카테고리") or "").strip()
        lv2 = str(r.get("Lv2코드") or "").strip()
        lv2_name = str(r.get("Lv2카테고리") or "").strip()
        if not lv1 or not lv2:
            continue
        if lv1 not in lv1_seen:
            lv1_seen[lv1] = lv1_name
            conn.execute(
                "INSERT INTO code_lv1(code, name, sort_order, use_yn) VALUES (?,?,?,?)",
                (lv1, lv1_name, len(lv1_seen), _yn(r.get("사용여부"))),
            )
        # Lv2는 전역 유일: 같은 코드면 같은 이름이어야 함
        if lv2 in lv2_seen:
            if lv2_seen[lv2] != lv2_name:
                errors.append(f"lv2 code conflict: {lv2} name '{lv2_seen[lv2]}' vs '{lv2_name}' (Lv1={lv1})")
            continue
        lv2_seen[lv2] = lv2_name
        conn.execute(
            "INSERT INTO code_lv2(code, name, sort_order, use_yn) VALUES (?,?,?,?)",
            (lv2, lv2_name, len(lv2_seen), _yn(r.get("사용여부"))),
        )

    from migrate_lv3_unique import next_lv3_code

    lv3_seen: set[str] = set()
    common_remap: dict[str, str] = {}
    for r in commons:
        common = str(r.get("공통지표코드") or "").strip().upper()
        lv1 = str(r.get("Lv1코드") or "").strip().upper()
        lv2 = str(r.get("Lv2코드") or "").strip()
        lv3 = str(r.get("Lv3코드") or "").strip()
        name = str(r.get("Lv3지표") or "").strip()
        unit = str(r.get("단위") or r.get("예상단위") or "").strip()
        common_yn = str(r.get("공통여부") or "단독").strip()
        definition_text = str(r.get("지표정의") or r.get("정의") or "").strip()
        calc_logic_text = str(r.get("산출로직") or "").strip()
        data_source = str(r.get("데이터원천") or r.get("원천") or r.get("원천상세") or "").strip()
        data_source_kind = str(r.get("데이터원천종류") or r.get("원천종류") or "").strip()
        calc_cycle = str(r.get("산출주기") or "").strip()
        calc_timing = str(r.get("산출시점") or "").strip()
        dept = str(r.get("주관부서") or r.get("담당부서") or r.get("부서") or "").strip()
        if not common:
            continue
        expected = f"{lv1}-{lv2}-{lv3}"
        if common != expected:
            errors.append(f"common mismatch: {common} != {expected}")
            continue
        orig_common = common
        if not lv3 or lv3 in lv3_seen:
            lv3 = next_lv3_code(conn)
            while lv3 in lv3_seen:
                n = int(lv3) + 1
                if n > 9999:
                    errors.append(f"lv3 exhausted while remapping {orig_common}")
                    lv3 = ""
                    break
                lv3 = f"{n:04d}"
            if not lv3:
                continue
            common = f"{lv1}-{lv2}-{lv3}"
        lv3_seen.add(lv3)
        common_remap[orig_common] = common
        try:
            from indicator_definition import normalize_data_source_kind
            data_source_kind = normalize_data_source_kind(data_source_kind)
            conn.execute(
                """INSERT INTO indicator_common
                   (common_code, lv1_code, lv2_code, lv3_code, name, unit, allowed_perf, common_yn, use_yn,
                    definition_text, calc_logic_text, dept, calc_cycle, calc_timing,
                    data_source_kind, data_source)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    common, lv1, lv2, lv3, name, unit, "", common_yn, "Y",
                    definition_text, calc_logic_text, dept, calc_cycle, calc_timing,
                    data_source_kind, data_source,
                ),
            )
        except Exception as e:
            errors.append(f"common insert {common}: {e}")

    for r in codes:
        indicator = str(r.get("지표코드") or "").strip().upper()
        common = str(r.get("공통지표코드") or "").strip().upper()
        lv1 = str(r.get("Lv1코드") or "").strip().upper()
        lv2 = str(r.get("Lv2코드") or "").strip()
        lv3 = str(r.get("Lv3코드") or "").strip()
        group = str(r.get("그룹코드") or "").strip().upper()
        perf = str(r.get("실적구분코드") or "").strip().upper()
        display = str(r.get("코드화지표명") or "").strip()
        legacy_unit = str(r.get("예상단위") or r.get("단위") or "").strip()
        detailed = str(r.get("상세지표정의") or r.get("상세정의") or "").strip()
        if not indicator:
            continue
        remapped = common_remap.get(common)
        if remapped and remapped != common:
            common = remapped
            parts = common.split("-")
            if len(parts) >= 3:
                lv1, lv2, lv3 = parts[0], parts[1], parts[2]
        try:
            composed = compose_indicator_code(lv1, lv2, lv3, perf, group)
        except ValueError as e:
            errors.append(f"compose fail {indicator}: {e}")
            continue
        if remapped is None and composed != indicator:
            errors.append(f"code mismatch: excel={indicator} composed={composed}")
            continue
        indicator = composed
        row = conn.execute(
            "SELECT name, unit FROM indicator_common WHERE common_code=?",
            (common,),
        ).fetchone()
        if not row:
            errors.append(f"missing common for {indicator}: {common}")
            continue
        unit = str(row["unit"] or "").strip() or legacy_unit
        if legacy_unit and not str(row["unit"] or "").strip():
            conn.execute(
                "UPDATE indicator_common SET unit=? WHERE common_code=? AND TRIM(COALESCE(unit,''))=''",
                (legacy_unit, common),
            )
            unit = legacy_unit
        try:
            conn.execute(
                """INSERT INTO indicator_code
                   (indicator_code, common_code, group_code, perf_code, display_name, unit, agg_type, use_yn,
                    detailed_definition_text)
                   VALUES (?,?,?,?,?,?,?,?,?)""",
                (
                    indicator, common, group, perf, display, unit, "", "Y",
                    detailed,
                ),
            )
        except Exception as e:
            errors.append(f"code insert {indicator}: {e}")

    if errors:
        conn.rollback()
        raise RuntimeError("import failed (%d errors):\n%s" % (len(errors), "\n".join(errors[:30])))

    from migrate_lv3_unique import ensure_lv3_unique_index
    ensure_lv3_unique_index(conn)
    conn.commit()
    result = {"ok": True, "counts": counts(conn), "errors": []}
    if own:
        conn.close()
    return result


def export_workbook(xlsx_path: Path | None = None, conn=None) -> Path:
    """현재 SQLite 코드마스터를 임포트와 동일 시트/컬럼 구조의 xlsx로 내보낸다."""
    from openpyxl import Workbook

    own = conn is None
    if own:
        conn = get_connection()
    path = Path(xlsx_path) if xlsx_path else Path(__file__).resolve().parent / "code_master_export.xlsx"

    try:
        wb = Workbook()

        ws = wb.active
        ws.title = SHEET_GROUP
        ws.append(["그룹코드", "그룹명", "정렬순서", "사용여부", "조직레벨", "상위코드"])
        for r in conn.execute(
            "SELECT code, name, sort_order, use_yn, org_level, parent_code FROM owner_group ORDER BY sort_order, code"
        ):
            ws.append([
                r["code"], r["name"], r["sort_order"],
                "사용" if r["use_yn"] == "Y" else "미사용",
                r["org_level"] or "GROUP",
                r["parent_code"] or "",
            ])

        ws = wb.create_sheet(SHEET_LV)
        ws.append(["Lv1코드", "Lv1카테고리", "Lv2코드", "Lv2카테고리", "사용여부"])
        # Lv1 · Lv2 독립 마스터를 카테시안이 아니라, 실제 사용 조합(indicator_common) + 미사용 마스터 보강
        lv1_rows = {
            r["code"]: r
            for r in conn.execute("SELECT code, name, use_yn FROM code_lv1 ORDER BY sort_order, code")
        }
        lv2_rows = {
            r["code"]: r
            for r in conn.execute("SELECT code, name, use_yn FROM code_lv2 ORDER BY sort_order, code")
        }
        pairs = conn.execute(
            """
            SELECT DISTINCT lv1_code, lv2_code FROM indicator_common
            ORDER BY lv1_code, lv2_code
            """
        ).fetchall()
        written = set()
        for r in pairs:
            l1 = lv1_rows.get(r["lv1_code"])
            l2 = lv2_rows.get(r["lv2_code"])
            if not l1 or not l2:
                continue
            written.add((r["lv1_code"], r["lv2_code"]))
            ws.append([
                r["lv1_code"], l1["name"], r["lv2_code"], l2["name"],
                "사용" if (l1["use_yn"] == "Y" and l2["use_yn"] == "Y") else "미사용",
            ])
        # 조합에 없는 Lv2도 마스터 보존용으로 첫 Lv1과 함께 export
        first_lv1 = next(iter(lv1_rows.values()), None)
        if first_lv1:
            for code, l2 in lv2_rows.items():
                if any(code == p[1] for p in written):
                    continue
                ws.append([
                    first_lv1["code"], first_lv1["name"], code, l2["name"],
                    "사용" if l2["use_yn"] == "Y" else "미사용",
                ])

        ws = wb.create_sheet(SHEET_COMMON)
        ws.append([
            "공통지표코드", "Lv1코드", "Lv2코드", "Lv3코드", "Lv3지표", "단위", "공통여부",
            "지표정의", "산출로직", "주관부서", "산출주기", "산출시점", "데이터원천종류", "데이터원천",
        ])
        for r in conn.execute(
            """
            SELECT common_code, lv1_code, lv2_code, lv3_code, name, unit, common_yn,
                   definition_text, calc_logic_text, dept, calc_cycle, calc_timing,
                   data_source_kind, data_source
            FROM indicator_common ORDER BY common_code
            """
        ):
            ws.append([
                r["common_code"], r["lv1_code"], r["lv2_code"], r["lv3_code"],
                r["name"], r["unit"] or "", r["common_yn"],
                r["definition_text"] or "", r["calc_logic_text"] or "", r["dept"] or "",
                r["calc_cycle"] or "", r["calc_timing"] or "",
                r["data_source_kind"] or "", r["data_source"] or "",
            ])

        ws = wb.create_sheet(SHEET_CODES)
        ws.append([
            "지표코드", "공통지표코드", "Lv1코드", "Lv2코드", "Lv3코드",
            "그룹코드", "실적구분코드", "코드화지표명", "상세지표정의",
        ])
        for r in conn.execute(
            """
            SELECT
              ic.indicator_code, ic.common_code,
              cm.lv1_code, cm.lv2_code, cm.lv3_code,
              ic.group_code, ic.perf_code, ic.display_name,
              ic.detailed_definition_text
            FROM indicator_code ic
            JOIN indicator_common cm ON cm.common_code = ic.common_code
            ORDER BY ic.indicator_code
            """
        ):
            ws.append([
                r["indicator_code"], r["common_code"],
                r["lv1_code"], r["lv2_code"], r["lv3_code"],
                r["group_code"], r["perf_code"], r["display_name"],
                r["detailed_definition_text"] or "",
            ])

        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        return path
    finally:
        if own:
            conn.close()


def main(argv=None):
    parser = argparse.ArgumentParser(description="Import KPI code master xlsx into SQLite")
    parser.add_argument("--xlsx", default=str(DEFAULT_XLSX), help="path to final xlsx")
    args = parser.parse_args(argv)
    try:
        result = import_workbook(Path(args.xlsx))
    except Exception as e:
        print("FAIL:", e, file=sys.stderr)
        return 1
    print("OK", result["counts"])
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
