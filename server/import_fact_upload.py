# -*- coding: utf-8 -*-
"""실적 엑셀 업로드: 평가월(YYYYMM) · 지표코드 · 실적 → 플랫폼 staging."""
from __future__ import annotations

import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from bank_export import push_to_bank
from fact_pipeline import INDICATOR_CODE_RE, parse_ym, refresh_facts

SHEET_NAME = "실적업로드"
TEMPLATE_PATH = Path(__file__).resolve().parent / "fact_upload_template.xlsx"
HEADERS = ["평가월", "지표코드", "실적"]

YM_RE = re.compile(r"^\d{6}$")


def normalize_eval_ym(value: Any) -> str:
    """평가월을 YYYYMM으로 정규화. 결과는 반드시 YYYYMM.

    허용 입력: 202606, 2026-06, 2026/06, 2026.6, datetime, Excel 숫자(202606.0)
    """
    if value is None or value == "":
        raise ValueError("평가월 없음")
    if hasattr(value, "strftime"):
        return value.strftime("%Y%m")
    if isinstance(value, bool):
        raise ValueError(f"평가월은 YYYYMM 형식이어야 합니다: {value!r}")
    if isinstance(value, (int, float)):
        if isinstance(value, float) and not value.is_integer():
            raise ValueError(f"평가월은 YYYYMM 형식이어야 합니다: {value!r}")
        s = str(int(value))
        if YM_RE.fullmatch(s):
            y, m = int(s[:4]), int(s[4:6])
            if m < 1 or m > 12:
                raise ValueError(f"평가월 월 범위 오류: {s}")
            return s
        raise ValueError(f"평가월은 YYYYMM 형식이어야 합니다: {value!r}")

    s = str(value).strip()
    if re.fullmatch(r"\d+\.0+", s):
        s = s.split(".", 1)[0]
    # YYYYMM
    if YM_RE.fullmatch(s):
        y, m = int(s[:4]), int(s[4:6])
        if m < 1 or m > 12:
            raise ValueError(f"평가월 월 범위 오류: {s}")
        return s
    # YYYY-MM / YYYY/MM / YYYY.MM (월 1~2자리)
    m = re.fullmatch(r"(\d{4})[-./](\d{1,2})", s)
    if m:
        y, mo = int(m.group(1)), int(m.group(2))
        if mo < 1 or mo > 12:
            raise ValueError(f"평가월 월 범위 오류: {s}")
        return f"{y}{mo:02d}"
    raise ValueError(f"평가월은 YYYYMM 형식이어야 합니다: {value!r}")


def group_from_indicator(code: str) -> str:
    parts = str(code or "").strip().upper().split("-")
    if len(parts) < 5:
        return ""
    return parts[-1]


def write_template(path: Path | None = None) -> Path:
    target = Path(path) if path else TEMPLATE_PATH
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(HEADERS)
    ws.append(["202606", "CAP-0001-0001-RAT-CIG", 0.21])
    ws.append(["202606", "CUS-0002-0002-NET-SG1", 12500])
    # guide sheet
    guide = wb.create_sheet("안내")
    guide.append(["컬럼", "설명"])
    guide.append(["평가월", "필수. YYYYMM (예: 202606)"])
    guide.append(["지표코드", "필수. Lv1-Lv2-Lv3-실적구분-그룹코드"])
    guide.append(["실적", "필수. 숫자 (기본단위)"])
    guide.append(["중복", "같은 평가월+지표코드는 파일 내 마지막 행 기준"])
    guide.append(["반영", "업로드 후 기존값과 다른 건을 미리보기 → 확인 시에만 DB 반영"])
    target.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target)
    return target


def _cell(row: tuple, idx: int):
    return row[idx] if idx < len(row) else None


def parse_fact_workbook(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    """엑셀 → (items, row_errors). 파일 내 중복(평가월+지표코드)은 마지막 행 승."""
    wb = load_workbook(path, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("빈 엑셀입니다")
    header = [str(c or "").strip() for c in rows[0]]
    col: dict[str, int] = {}
    aliases = {
        "평가월": ("평가월", "eval_ym", "evalYm", "연월", "YM"),
        "지표코드": ("지표코드", "indicator_code", "indicatorCode", "코드"),
        "실적": ("실적", "actual", "실적값", "값"),
    }
    for field, names in aliases.items():
        for i, h in enumerate(header):
            if h in names:
                col[field] = i
                break

    if set(col) == {"평가월", "지표코드", "실적"}:
        data_rows = rows[1:]
    else:
        # 헤더 없이 3열 데이터로 시작하는지 시도
        try:
            normalize_eval_ym(_cell(rows[0], 0))
            str(_cell(rows[0], 1) or "").strip()
            float(_cell(rows[0], 2))
            col = {"평가월": 0, "지표코드": 1, "실적": 2}
            data_rows = rows
        except Exception:
            missing = [k for k in ("평가월", "지표코드", "실적") if k not in col]
            raise ValueError(f"필수 컬럼 없음: {', '.join(missing) or '평가월, 지표코드, 실적'}") from None

    by_key: dict[tuple[str, str], dict] = {}
    errors: list[str] = []
    for ridx, row in enumerate(data_rows, start=2 if data_rows is not rows else 1):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        try:
            ym = normalize_eval_ym(_cell(row, col["평가월"]))
            code = str(_cell(row, col["지표코드"]) or "").strip().upper()
            if not code:
                raise ValueError("지표코드 없음")
            if not INDICATOR_CODE_RE.match(code):
                raise ValueError(f"지표코드 형식 오류: {code}")
            raw_actual = _cell(row, col["실적"])
            if raw_actual is None or str(raw_actual).strip() == "":
                raise ValueError("실적 없음")
            actual = float(raw_actual)
            group = group_from_indicator(code)
            by_key[(ym, code)] = {
                "eval_ym": ym,
                "indicator_code": code,
                "actual": actual,
                "group_code": group,
                "row_no": ridx,
            }
        except Exception as e:
            errors.append(f"행 {ridx}: {e}")
    if errors and not by_key:
        raise ValueError("유효한 행이 없습니다.\n" + "\n".join(errors[:20]))
    items = list(by_key.values())
    items.sort(key=lambda x: (x["eval_ym"], x["indicator_code"]))
    return items, errors


def _actuals_equal(a: Any, b: Any) -> bool:
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) < 1e-9
    except (TypeError, ValueError):
        return False


def _write_change_log(
    conn,
    *,
    batch_id: int,
    eval_ym: str,
    indicator_code: str,
    group_code: str,
    prev_actual: Any,
    new_actual: Any,
    change_kind: str,
    action: str,
    acted_by: str,
) -> None:
    conn.execute(
        """
        INSERT INTO fact_upload_change_log(
          batch_id, eval_ym, indicator_code, group_code,
          prev_actual, new_actual, change_kind, action, acted_by
        ) VALUES (?,?,?,?,?,?,?,?,?)
        """,
        (
            batch_id,
            eval_ym,
            indicator_code,
            group_code or "",
            prev_actual,
            new_actual,
            change_kind,
            action,
            acted_by,
        ),
    )


def _item_payload(it: dict) -> dict:
    return {
        "row_no": it.get("row_no"),
        "eval_ym": it["eval_ym"],
        "indicator_code": it["indicator_code"],
        "group_code": it.get("group_code") or "",
        "prev_actual": it.get("prev_actual"),
        "actual": it.get("actual"),
        "change_kind": it.get("change_kind") or "",
        "status": it.get("status") or "ok",
        "error_text": it.get("error_text") or "",
    }


def preview_fact_upload(
    conn,
    path: Path,
    *,
    filename: str = "",
    uploaded_by: str = "ui",
) -> dict[str, Any]:
    """엑셀 파싱·검증·기존값 비교. DB 실적(fact_collect)은 아직 변경하지 않음."""
    items, row_errors = parse_fact_workbook(path)
    started = datetime.now(timezone.utc).isoformat()
    cur = conn.execute(
        """
        INSERT INTO fact_upload_batch(
          filename, status, counts_json, error_text, uploaded_by, created_at
        ) VALUES (?,?,?,?,?,?)
        """,
        (filename or path.name, "preview", "{}", "", uploaded_by, started),
    )
    batch_id = cur.lastrowid

    known = {
        r["indicator_code"]
        for r in conn.execute("SELECT indicator_code FROM indicator_code WHERE use_yn='Y'").fetchall()
    }

    ok_items: list[dict] = []
    item_errors: list[str] = list(row_errors)
    changed: list[dict] = []
    new_rows: list[dict] = []
    same_rows: list[dict] = []

    for it in items:
        if it["indicator_code"] not in known:
            msg = f"미등록 지표코드 {it['indicator_code']}"
            item_errors.append(f"행 {it['row_no']}: {msg}")
            conn.execute(
                """
                INSERT INTO fact_upload_item(
                  batch_id, eval_ym, indicator_code, group_code, actual, prev_actual,
                  change_kind, status, error_text, export_status, row_no
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    batch_id, it["eval_ym"], it["indicator_code"], it["group_code"], it["actual"], None,
                    "error", "error", msg, "skipped", it["row_no"],
                ),
            )
            _write_change_log(
                conn,
                batch_id=batch_id,
                eval_ym=it["eval_ym"],
                indicator_code=it["indicator_code"],
                group_code=it["group_code"],
                prev_actual=None,
                new_actual=it["actual"],
                change_kind="error",
                action="preview",
                acted_by=uploaded_by,
            )
            continue

        existing = conn.execute(
            "SELECT actual FROM fact_collect WHERE eval_ym=? AND indicator_code=?",
            (it["eval_ym"], it["indicator_code"]),
        ).fetchone()
        prev = float(existing["actual"]) if existing and existing["actual"] is not None else None
        if existing is None:
            kind = "new"
        elif _actuals_equal(prev, it["actual"]):
            kind = "same"
        else:
            kind = "changed"

        row = {
            **it,
            "prev_actual": prev,
            "change_kind": kind,
            "status": "ok",
            "error_text": "",
        }
        ok_items.append(row)
        conn.execute(
            """
            INSERT INTO fact_upload_item(
              batch_id, eval_ym, indicator_code, group_code, actual, prev_actual,
              change_kind, status, error_text, export_status, row_no
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                batch_id, row["eval_ym"], row["indicator_code"], row["group_code"], row["actual"],
                prev, kind, "ok", "", "held", row["row_no"],
            ),
        )
        _write_change_log(
            conn,
            batch_id=batch_id,
            eval_ym=row["eval_ym"],
            indicator_code=row["indicator_code"],
            group_code=row["group_code"],
            prev_actual=prev,
            new_actual=row["actual"],
            change_kind=kind,
            action="preview",
            acted_by=uploaded_by,
        )
        payload = _item_payload(row)
        if kind == "changed":
            changed.append(payload)
        elif kind == "new":
            new_rows.append(payload)
        else:
            same_rows.append(payload)

    status = "preview" if ok_items else "error"
    counts = {
        "rows_ok": len(ok_items),
        "rows_new": len(new_rows),
        "rows_changed": len(changed),
        "rows_same": len(same_rows),
        "rows_error": len(item_errors),
        "yms": sorted({it["eval_ym"] for it in ok_items}),
    }
    err_text = "\n".join(item_errors[:50])
    conn.execute(
        """
        UPDATE fact_upload_batch
        SET status=?, counts_json=?, error_text=?
        WHERE id=?
        """,
        (status, json.dumps(counts, ensure_ascii=False), err_text, batch_id),
    )
    conn.commit()
    return {
        "ok": bool(ok_items),
        "batch_id": batch_id,
        "status": status,
        "counts": counts,
        "changed": changed,
        "new": new_rows,
        "same": same_rows[:50],
        "errors": item_errors[:50],
        "needs_confirm": bool(ok_items),
    }


def _apply_collect_items(
    conn,
    batch_id: int,
    ok_items: list[dict],
    *,
    started: str,
    refresh: bool = True,
) -> tuple[list[str], list[str]]:
    """fact_collect 반영 + (옵션) 산출/달성 재계산. (refreshed, errors)"""
    refreshed: list[str] = []
    errors: list[str] = []
    if not ok_items:
        return refreshed, errors

    sync_by_ym: dict[str, int] = {}
    for ym in sorted({it["eval_ym"] for it in ok_items}):
        cur_sb = conn.execute(
            "INSERT INTO sync_batch(mode, eval_ym, status, started_at) VALUES (?,?,?,?)",
            ("upload", ym, "running", started),
        )
        sync_by_ym[ym] = cur_sb.lastrowid

    for it in ok_items:
        conn.execute(
            """
            INSERT INTO fact_collect(eval_ym, indicator_code, actual, batch_id, fetched_at)
            VALUES (?,?,?,?,?)
            ON CONFLICT(eval_ym, indicator_code) DO UPDATE SET
              actual=excluded.actual,
              batch_id=excluded.batch_id,
              fetched_at=excluded.fetched_at
            """,
            (
                it["eval_ym"],
                it["indicator_code"],
                it["actual"],
                sync_by_ym[it["eval_ym"]],
                started,
            ),
        )

    for ym, sb_id in sync_by_ym.items():
        conn.execute(
            """
            UPDATE sync_batch
            SET status=?, counts_json=?, finished_at=?
            WHERE id=?
            """,
            (
                "ok",
                json.dumps(
                    {
                        "source": "fact_upload",
                        "upload_batch_id": batch_id,
                        "rows": sum(1 for x in ok_items if x["eval_ym"] == ym),
                    },
                    ensure_ascii=False,
                ),
                datetime.now(timezone.utc).isoformat(),
                sb_id,
            ),
        )

    if refresh:
        for ym in sorted(sync_by_ym):
            year, month = parse_ym(ym)
            try:
                refresh_facts(conn, year, month, skip_collect=True)
                refreshed.append(ym)
            except Exception as e:
                errors.append(f"refresh {ym} 실패: {e}")
    return refreshed, errors


def confirm_fact_upload(
    conn,
    batch_id: int,
    *,
    acted_by: str = "ui",
    refresh: bool = True,
) -> dict[str, Any]:
    """미리보기 배치를 확인 후 fact_collect에 반영."""
    batch = conn.execute(
        "SELECT * FROM fact_upload_batch WHERE id=?", (int(batch_id),)
    ).fetchone()
    if not batch:
        raise ValueError("업로드 배치를 찾을 수 없습니다")
    if batch["status"] != "preview":
        raise ValueError(f"확인 가능한 상태가 아닙니다 (현재: {batch['status']})")

    rows = conn.execute(
        """
        SELECT * FROM fact_upload_item
        WHERE batch_id=? AND status='ok'
        ORDER BY eval_ym, indicator_code
        """,
        (int(batch_id),),
    ).fetchall()
    ok_items = [dict(r) for r in rows]
    if not ok_items:
        raise ValueError("반영할 유효 행이 없습니다")

    started = datetime.now(timezone.utc).isoformat()
    refreshed, apply_errors = _apply_collect_items(
        conn, int(batch_id), ok_items, started=started, refresh=refresh,
    )

    for it in ok_items:
        conn.execute(
            "UPDATE fact_upload_item SET export_status='pending' WHERE id=?",
            (it["id"],),
        )
        _write_change_log(
            conn,
            batch_id=int(batch_id),
            eval_ym=it["eval_ym"],
            indicator_code=it["indicator_code"],
            group_code=it.get("group_code") or "",
            prev_actual=it.get("prev_actual"),
            new_actual=it.get("actual"),
            change_kind=it.get("change_kind") or "",
            action="confirm",
            acted_by=acted_by,
        )

    status = "ok" if not apply_errors else "partial"
    try:
        prev_counts = json.loads(batch["counts_json"] or "{}")
    except Exception:
        prev_counts = {}
    counts = {
        **prev_counts,
        "rows_ok": len(ok_items),
        "rows_applied": len(ok_items),
        "refreshed": refreshed,
        "confirmed_at": started,
        "confirmed_by": acted_by,
    }
    err_text = "\n".join(
        [x for x in ((batch["error_text"] or "").split("\n") if batch["error_text"] else []) if x]
        + apply_errors
    )[:4000]
    conn.execute(
        """
        UPDATE fact_upload_batch
        SET status=?, counts_json=?, error_text=?, finished_at=?
        WHERE id=?
        """,
        (status, json.dumps(counts, ensure_ascii=False), err_text, started, int(batch_id)),
    )
    conn.commit()
    return {
        "ok": True,
        "batch_id": int(batch_id),
        "status": status,
        "counts": counts,
        "errors": apply_errors,
    }


def cancel_fact_upload(
    conn,
    batch_id: int,
    *,
    acted_by: str = "ui",
) -> dict[str, Any]:
    """미리보기 배치 취소 (DB 실적 미반영)."""
    batch = conn.execute(
        "SELECT * FROM fact_upload_batch WHERE id=?", (int(batch_id),)
    ).fetchone()
    if not batch:
        raise ValueError("업로드 배치를 찾을 수 없습니다")
    if batch["status"] != "preview":
        raise ValueError(f"취소 가능한 상태가 아닙니다 (현재: {batch['status']})")

    now = datetime.now(timezone.utc).isoformat()
    items = conn.execute(
        "SELECT * FROM fact_upload_item WHERE batch_id=?", (int(batch_id),)
    ).fetchall()
    for it in items:
        conn.execute(
            "UPDATE fact_upload_item SET export_status='skipped' WHERE id=?",
            (it["id"],),
        )
        _write_change_log(
            conn,
            batch_id=int(batch_id),
            eval_ym=it["eval_ym"] or "",
            indicator_code=it["indicator_code"] or "",
            group_code=it["group_code"] or "",
            prev_actual=it["prev_actual"] if "prev_actual" in it.keys() else None,
            new_actual=it["actual"],
            change_kind=it["change_kind"] if "change_kind" in it.keys() else "",
            action="cancel",
            acted_by=acted_by,
        )

    try:
        counts = json.loads(batch["counts_json"] or "{}")
    except Exception:
        counts = {}
    counts["cancelled_by"] = acted_by
    counts["cancelled_at"] = now
    conn.execute(
        """
        UPDATE fact_upload_batch
        SET status='cancelled', counts_json=?, finished_at=?
        WHERE id=?
        """,
        (json.dumps(counts, ensure_ascii=False), now, int(batch_id)),
    )
    conn.commit()
    return {"ok": True, "batch_id": int(batch_id), "status": "cancelled", "counts": counts}


def import_fact_upload(
    conn,
    path: Path,
    *,
    filename: str = "",
    uploaded_by: str = "ui",
    apply_collect: bool = True,
    refresh: bool = True,
) -> dict[str, Any]:
    """하위호환: apply_collect=False면 preview만, True면 preview 직후 즉시 confirm."""
    preview = preview_fact_upload(
        conn, path, filename=filename, uploaded_by=uploaded_by,
    )
    if not apply_collect or not preview.get("ok"):
        return preview
    return confirm_fact_upload(
        conn, preview["batch_id"], acted_by=uploaded_by, refresh=refresh,
    )


def list_upload_batches(conn, *, limit: int = 50) -> list[dict]:
    rows = conn.execute(
        """
        SELECT * FROM fact_upload_batch
        ORDER BY id DESC
        LIMIT ?
        """,
        (int(limit),),
    ).fetchall()
    return [dict(r) for r in rows]


def list_upload_items(conn, batch_id: int) -> list[dict]:
    rows = conn.execute(
        """
        SELECT * FROM fact_upload_item
        WHERE batch_id=?
        ORDER BY
          CASE change_kind WHEN 'changed' THEN 0 WHEN 'new' THEN 1 WHEN 'same' THEN 2 ELSE 3 END,
          eval_ym, indicator_code
        """,
        (int(batch_id),),
    ).fetchall()
    return [dict(r) for r in rows]


def list_upload_change_logs(conn, batch_id: int | None = None, *, limit: int = 200) -> list[dict]:
    if batch_id is not None:
        rows = conn.execute(
            """
            SELECT * FROM fact_upload_change_log
            WHERE batch_id=?
            ORDER BY id DESC
            LIMIT ?
            """,
            (int(batch_id), int(limit)),
        ).fetchall()
    else:
        rows = conn.execute(
            """
            SELECT * FROM fact_upload_change_log
            ORDER BY id DESC
            LIMIT ?
            """,
            (int(limit),),
        ).fetchall()
    return [dict(r) for r in rows]


def export_pending_uploads_to_bank(conn, *, triggered_by: str = "scheduler") -> dict[str, Any]:
    """자정 배치용: 입력 완료(pending) 실적을 행내 DB로 전송.

    Freeze 여부와 무관하게 pending을 전송한다.
    Freeze는 수정 잠금(최종 마감)용이며 전송 게이트가 아니다.
    """
    started = datetime.now(timezone.utc).isoformat()
    pending = conn.execute(
        """
        SELECT fui.*
        FROM fact_upload_item fui
        WHERE fui.status='ok' AND fui.export_status='pending'
        ORDER BY fui.eval_ym, fui.indicator_code
        """
    ).fetchall()
    items = [dict(r) for r in pending]
    if not items:
        return {"ok": True, "exported": 0, "message": "전송 대기(pending) 실적 없음"}

    # 연월별로 bank_export_batch 기록 (어댑터 호출)
    by_ym: dict[str, list[dict]] = {}
    for it in items:
        by_ym.setdefault(it["eval_ym"], []).append(it)

    exported = 0
    batch_ids = []
    for ym, rows in by_ym.items():
        cur = conn.execute(
            """
            INSERT INTO bank_export_batch(eval_ym, status, triggered_by, counts_json, error_text, started_at)
            VALUES (?,?,?,?,?,?)
            """,
            (ym, "running", triggered_by, "{}", "", started),
        )
        be_id = cur.lastrowid
        batch_ids.append(be_id)
        payload_items = []
        for r in rows:
            conn.execute(
                """
                INSERT INTO bank_export_item(
                  batch_id, eval_ym, group_code, indicator_code, actual, calc_kind,
                  monthly_target, converted_achievement, payload_json
                ) VALUES (?,?,?,?,?,?,?,?,?)
                """,
                (
                    be_id, ym, r.get("group_code") or "", r["indicator_code"], r["actual"], "UPLOAD",
                    None, None,
                    json.dumps(
                        {"source": "fact_upload", "upload_item_id": r["id"], "upload_batch_id": r["batch_id"]},
                        ensure_ascii=False,
                    ),
                ),
            )
            payload_items.append({
                "eval_ym": ym,
                "group_code": r.get("group_code"),
                "indicator_code": r["indicator_code"],
                "actual": r["actual"],
                "source": "fact_upload",
            })
            conn.execute(
                "UPDATE fact_upload_item SET export_status='exported', exported_at=? WHERE id=?",
                (datetime.now(timezone.utc).isoformat(), r["id"]),
            )
            exported += 1
        try:
            push_to_bank({"id": be_id, "eval_ym": ym, "triggered_by": triggered_by}, payload_items)
            conn.execute(
                """
                UPDATE bank_export_batch
                SET status='ok', counts_json=?, finished_at=?
                WHERE id=?
                """,
                (
                    json.dumps({"items": len(rows), "source": "fact_upload"}, ensure_ascii=False),
                    datetime.now(timezone.utc).isoformat(),
                    be_id,
                ),
            )
        except Exception as e:
            conn.execute(
                """
                UPDATE bank_export_batch
                SET status='error', error_text=?, finished_at=?
                WHERE id=?
                """,
                (str(e), datetime.now(timezone.utc).isoformat(), be_id),
            )
            for r in rows:
                conn.execute(
                    "UPDATE fact_upload_item SET export_status='failed' WHERE id=?",
                    (r["id"],),
                )
    conn.commit()
    return {"ok": True, "exported": exported, "bank_export_batch_ids": batch_ids, "yms": list(by_ym.keys())}
