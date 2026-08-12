# -*- coding: utf-8 -*-
"""05_평가체계맵핑 시트 기반으로 2026년 가상 평가배치 생성."""
from __future__ import annotations

import math
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

from db import get_connection
from kpi_api import Handler

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "은행_KPI_3단분류_코드마스터_최종.xlsx"
YEAR = 2026
EFFECTIVE_MONTH = 1

# 평가배치 Lv1 (코드마스터 Lv1과 별개)
EVAL_LV1_PROFIT = "본원적 수익력"
EVAL_LV1_SOUND = "건전성"
EVAL_LV1_CUST = "고객"
EVAL_LV1_CONNECT = "연결과 확장"
EVAL_LV1_DIGITAL = "디지털"  # 전행(SHB)만

# 코드마스터 Lv1코드 → 평가 Lv1
CODE_LV1_TO_EVAL = {
    "PFT": EVAL_LV1_PROFIT,
    "COS": EVAL_LV1_PROFIT,
    "LON": EVAL_LV1_PROFIT,
    "YUD": EVAL_LV1_PROFIT,
    "IBF": EVAL_LV1_PROFIT,
    "CMT": EVAL_LV1_PROFIT,
    "PRD": EVAL_LV1_PROFIT,
    "FXM": EVAL_LV1_PROFIT,
    "GLB": EVAL_LV1_PROFIT,
    "WMT": EVAL_LV1_PROFIT,
    "CAP": EVAL_LV1_SOUND,
    "CUS": EVAL_LV1_CUST,
    "CEX": EVAL_LV1_CUST,
    "SFM": EVAL_LV1_CUST,
    "DGP": EVAL_LV1_CONNECT,
    "INP": EVAL_LV1_CONNECT,
    "RPN": EVAL_LV1_CONNECT,
    "STG": EVAL_LV1_CONNECT,
}


def map_eval_lv1(code_lv1: str, group_code: str) -> str:
    base = CODE_LV1_TO_EVAL.get((code_lv1 or "").strip().upper(), EVAL_LV1_CONNECT)
    if group_code == "SHB" and base == EVAL_LV1_CONNECT:
        return EVAL_LV1_DIGITAL
    return base


def normalize_eval_hierarchy(label: str, map_lv2: str, map_lv3: str, eval_lv1: str) -> tuple[str, str]:
    """코드마스터 정규화지표(거의 Label과 1:1)를 평가용 Lv2/Lv3 버킷으로 재분류.

    - Label(=표시명)은 리프 유일키로 유지
    - Lv3는 주제 버킷이라 한 Lv3 아래 여러 Label이 붙도록 함
    """
    text = f"{label or ''} {map_lv3 or ''}"
    lv2 = (map_lv2 or "").strip() or "미분류"

    # Lv2가 Label과 동일(리프 복제)일 때만 폴백. 단순 부분문자열 일치는 정상(예: 기관고객 ⊂ 기관고객 순이자이익)
    compact_label = (label or "").replace(" ", "")
    compact_lv2 = lv2.replace(" ", "")
    if compact_lv2 and compact_label and compact_lv2 == compact_label:
        lv2 = {
            EVAL_LV1_PROFIT: "수익성",
            EVAL_LV1_SOUND: "리스크",
            EVAL_LV1_CUST: "고객기반",
            EVAL_LV1_CONNECT: "사업확장",
            EVAL_LV1_DIGITAL: "디지털",
        }.get(eval_lv1, "기타")

    if any(k in text for k in ("예금", "수신", "저원가", "조달비용", "평잔")) and "여신" not in text:
        lv3 = "예금"
    elif any(k in text for k in ("여신", "대출", "ROC", "RORWA", "우량여신")):
        lv3 = "여신"
    elif any(k in text for k in ("세전이익", "영업이익", "순이자", "비이자", "수수료이익", "관리손익", "수수료")):
        lv3 = "손익"
    elif any(k in text for k in ("계약", "갱신", "수주")):
        lv3 = "계약"
    elif any(k in text for k in ("ROI", "제휴사업")):
        lv3 = "제휴성과"
    elif any(k in text for k in ("유지율", "전환율", "이탈", "활동성", "주거래")):
        lv3 = "관계유지"
    elif any(k in text for k in ("신규", "유치", "유입", "순증", "고객 수", "고객수")):
        lv3 = "신규·기반"
    elif any(k in text for k in ("점수", "경험", "만족", "NPS", "민원")):
        lv3 = "고객경험"
    elif any(k in text for k in ("연체", "부실", "NPL", "건전", "충당금")):
        lv3 = "건전성지표"
    elif any(k in text for k in ("디지털", "앱", "플랫폼", "MAU", "비대면", "오픈뱅킹")):
        lv3 = "디지털이용"
    elif any(k in text for k in ("비용률", "경비율", "생산성", "1인당", "NIM")):
        lv3 = "효율"
    else:
        # 맵 Lv3가 Label과 다르고 짧으면 유지, 아니면 기타
        cand = (map_lv3 or "").strip()
        if cand and cand.replace(" ", "") not in compact_label and len(cand) <= 12:
            lv3 = cand
        else:
            lv3 = "기타"

    return lv2, lv3


def _equal_weights(n: int) -> list[float]:
    """그룹 내 균등 비중. 소수 둘째자리, 합계 100.00."""
    if n <= 0:
        return []
    base = math.floor(10000 / n) / 100
    weights = [base] * n
    rem = round(100.0 - sum(weights), 2)
    i = 0
    while rem >= 0.009:
        weights[i] = round(weights[i] + 0.01, 2)
        rem = round(rem - 0.01, 2)
        i = (i + 1) % n
    drift = round(100.0 - sum(weights), 2)
    weights[-1] = round(weights[-1] + drift, 2)
    return weights


def _guess_unit(perf_code: str, label: str, lv3: str = "") -> str:
    """기본단위만 사용. 만/억/조는 표시단에서 자동 축약."""
    text = f"{label} {lv3}"
    p = (perf_code or "").upper()
    if p == "RAT" or any(k in text for k in ("률", "율", "비중", "ROC", "RORWA", "NIM", "점수")):
        return "%"
    if any(k in text for k in ("고객", "회원", "MAU", "이용자")) and p in ("NET", "NEW", "OUT", "ETC", ""):
        if "손익" in text or "이익" in text or "수익" in text:
            return "원"
        return "명"
    if any(k in text for k in ("이익", "손익", "예금", "여신", "평잔", "잔액", "지원", "취급")):
        return "원"
    if p in ("NET", "NEW", "OUT"):
        return "원"
    return "건"


def _guess_mode(perf_code: str, label: str, unit: str) -> str:
    text = label or ""
    p = (perf_code or "").upper()
    if unit == "%" or p == "RAT" or any(k in text for k in ("률", "율", "비중", "ROC", "RORWA", "NIM", "점수")):
        return "flat"
    return "linear"


def _guess_goal_direction(label: str) -> str:
    text = label or ""
    # 낮을수록 좋은 지표
    if any(k in text for k in ("비용률", "경비율", "연체", "부실", "NPL", "이탈", "민원")):
        return "decrease"
    return "increase"


def _scale(group_code: str, bank: float, group: float) -> float:
    return float(bank if group_code == "SHB" else group)


def _guess_annual_target(group_code: str, code_lv1: str, perf_code: str, label: str, unit: str) -> float:
    text = label or ""
    p = (perf_code or "").upper()
    lv1 = (code_lv1 or "").upper()

    # 비율·점수류
    if unit == "%" or p == "RAT" or any(k in text for k in ("률", "율", "비중", "ROC", "RORWA", "NIM", "점수")):
        if "준수" in text:
            return 100.0
        if "유지율" in text:
            return 82.0
        if "비중" in text and "저원가" in text:
            return 48.0
        if "비중" in text:
            return 35.0
        if "NIM" in text.upper():
            return 1.85
        if "RORWA" in text.upper():
            return 1.60
        if "ROC" in text.upper():
            return 1.80
        if "점수" in text or "경험" in text:
            return 85.0
        if "생산성" in text or "1인당" in text or "고객당" in text:
            return 120.0  # 지수/상대값 가정
        if any(k in text for k in ("비용률", "경비율")):
            return 45.0
        if any(k in text for k in ("연체", "부실", "NPL")):
            return 0.80
        return 100.0

    # 고객 수
    if unit == "명":
        if "주거래" in text:
            return _scale(group_code, 120000, 18000)
        if "활동" in text or "MAU" in text.upper():
            return _scale(group_code, 350000, 45000)
        if "신규" in text:
            return _scale(group_code, 80000, 12000)
        return _scale(group_code, 50000, 8000)

    # 손익·잔액 — 저장은 원 단위 (표시단에서 만/억/조)
    won = 100_000_000.0  # 기존 억원 스케일 가정값 → 원
    if unit == "원" or any(k in text for k in ("이익", "손익", "예금", "여신", "평잔", "잔액")):
        if "세전이익" in text or "관리세전" in text:
            return _scale(group_code, 18000, 2200) * won
        if "충당금" in text and "영업이익" in text:
            return _scale(group_code, 22000, 2600) * won
        if "순이자" in text:
            return _scale(group_code, 14000, 1600) * won
        if "비이자" in text:
            return _scale(group_code, 4500, 520) * won
        if "핵심예금" in text:
            return _scale(group_code, 25000, 3200) * won
        if "우량여신" in text or ("여신" in text and "순증" in text):
            return _scale(group_code, 18000, 2400) * won
        if "서민" in text or "지원" in text:
            return _scale(group_code, 8000, 900) * won
        if lv1 == "IBF":
            return _scale(group_code, 3500, 2800) * won
        if lv1 == "CMT":
            return _scale(group_code, 2800, 2200) * won
        if lv1 == "WMT":
            return _scale(group_code, 4000, 1500) * won
        if lv1 == "GLB":
            return _scale(group_code, 3000, 1800) * won
        return _scale(group_code, 5000, 800) * won

    # 플랫폼·제휴 등
    if lv1 in ("DGP", "INP", "RPN", "STG"):
        if "건" in unit or unit == "건":
            return _scale(group_code, 120000, 25000)
        return _scale(group_code, 10000, 2500)

    return _scale(group_code, 1000, 100)


def _guess_baseline(mode: str, annual: float, label: str) -> float:
    if mode != "linear":
        return 0.0
    # 연초 기준실적: 연간목표의 약 70~80% 누적 시작점 가정
    ratio = 0.72 if "순증" in (label or "") else 0.78
    return round(annual * ratio, 2)


def load_map_rows() -> list[dict]:
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[5]]  # 05_평가체계맵핑
    rows = list(ws.iter_rows(values_only=True))
    out = []
    for r in rows[1:]:
        if not r or not r[14]:
            continue
        kind = str(r[4] or "").strip()
        if kind != "결과":
            continue
        out.append({
            "group_name": str(r[0] or "").strip(),
            "group_code": str(r[1] or "").strip().upper(),
            "sort_order": int(r[2] or 0),
            "label": str(r[3] or r[11] or "").strip(),
            "code_lv1": str(r[6] or "").strip().upper(),
            "code_lv1_name": str(r[7] or "").strip(),
            "eval_lv2": str(r[9] or "").strip(),
            "eval_lv3": str(r[11] or "").strip(),
            "perf_code": str(r[12] or "").strip().upper(),
            "indicator_code": str(r[14] or "").strip().upper(),
        })
    return out


def build_items() -> list[dict]:
    with get_connection() as conn:
        db_codes = {
            r["indicator_code"]: dict(r)
            for r in conn.execute(
                "SELECT indicator_code, group_code, display_name, unit FROM indicator_code"
            )
        }

    by_group: dict[str, list] = defaultdict(list)
    skipped = []
    for row in load_map_rows():
        code = row["indicator_code"]
        if code not in db_codes:
            skipped.append(code)
            continue
        if db_codes[code]["group_code"] != row["group_code"]:
            skipped.append(f"{code}:group_mismatch")
            continue
        by_group[row["group_code"]].append(row)

    items = []
    for group_code, rows in sorted(by_group.items()):
        rows = sorted(rows, key=lambda x: (x["sort_order"], x["indicator_code"]))
        weights = _equal_weights(len(rows))
        for i, row in enumerate(rows):
            meta = db_codes[row["indicator_code"]]
            unit = (meta.get("unit") or "").strip() or _guess_unit(row["perf_code"], row["label"], row["eval_lv3"])
            mode = _guess_mode(row["perf_code"], row["label"], unit)
            annual = _guess_annual_target(group_code, row["code_lv1"], row["perf_code"], row["label"], unit)
            baseline = _guess_baseline(mode, annual, row["label"])
            direction = _guess_goal_direction(row["label"])
            eval_lv1 = map_eval_lv1(row["code_lv1"], group_code)
            label = row["label"] or meta.get("display_name") or row["indicator_code"]
            eval_lv2, eval_lv3 = normalize_eval_hierarchy(
                label, row["eval_lv2"], row["eval_lv3"], eval_lv1,
            )
            items.append({
                "indicator_code": row["indicator_code"],
                "group_code": group_code,
                "mgmt_tool": "KPI",
                "eval_category_lv1": eval_lv1,
                "eval_category_lv2": eval_lv2,
                "eval_category_lv3": eval_lv3,
                "label": label,
                "unit": unit,
                "weight": round(float(weights[i]), 2),
                "is_core": "Y" if any(k in (label or "") for k in ("세전이익", "조정 ROC", "핵심예금", "주거래", "NPL", "연체")) else "N",
                "annual_target": round(float(annual), 2),
                "baseline_actual": round(float(baseline), 2),
                "achievement_mode": mode,
                "goal_direction": direction,
                "score_rule": "1",
                "penalty_rule": "1",
                "cap_max": 130,
                "cap_min": 0,
                "remark": f"05맵핑·결과지표 / 코드Lv1={row['code_lv1'] or row['code_lv1_name']}",
                "sort_order": row["sort_order"] or (i + 1),
                "use_yn": "Y",
            })

    return items, skipped, {g: len(v) for g, v in by_group.items()}


def main():
    if not XLSX.exists():
        raise SystemExit(f"missing workbook: {XLSX}")
    items, skipped, group_counts = build_items()
    print(f"groups={group_counts}")
    print(f"items={len(items)} skipped={len(skipped)}")
    if skipped:
        print("skipped sample:", skipped[:10])

    handler = Handler.__new__(Handler)
    result = handler._replace_eval_configs(
        YEAR,
        EFFECTIVE_MONTH,
        items,
        "05_평가체계맵핑 기반 가상 2026 평가체계 (Lv2/Lv3 버킷·비중0.01·목표보정)",
    )
    print(
        f"saved year={YEAR} month={EFFECTIVE_MONTH} "
        f"plan_set_id={result.get('plan_set_id')} "
        f"resolved_items={len(result.get('items') or [])}"
    )


if __name__ == "__main__":
    main()
