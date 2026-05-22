# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, DataBarRule
import os, random

# ── 그룹별 KPI 정의 (8그룹 x 20개) ──────────────────────────────

GROUPS = {
    "영업추진1그룹": {
        "code": "SL1", "dept": "영업추진1부",
        "kpis": [
            ("PWM 10억 고객수", "명", "총량", "KPI", 8),
            ("PWM 100억 고객수", "명", "총량", "KPI", 8),
            ("PWM 관리자산 총액", "억원", "총량", "KPI", 7),
            ("PB 고객 순증", "명", "연간신규", "KPI", 7),
            ("대출 실행액 (기업)", "억원", "연간신규", "KPI", 6),
            ("대출 실행액 (가계)", "억원", "연간신규", "KPI", 6),
            ("수신 평잔 증감", "억원", "총량", "KPI", 5),
            ("비이자수익", "억원", "연간신규", "KPI", 5),
            ("신규 VIP 고객 유치", "명", "연간신규", "전략과제", 5),
            ("고객 이탈률", "%", "비율", "KPI", 5),
            ("교차판매 비율", "%", "비율", "전략과제", 4),
            ("방카슈랑스 판매액", "억원", "연간신규", "전략과제", 4),
            ("펀드 판매잔액", "억원", "총량", "전략과제", 4),
            ("ELS 판매액", "억원", "연간신규", "전략과제", 3),
            ("고객 만족도 점수", "점", "비율", "모니터링", 3),
            ("영업점 방문 고객수", "명", "총량", "모니터링", 3),
            ("디지털 채널 전환율", "%", "비율", "모니터링", 3),
            ("직원 1인당 수익", "백만원", "비율", "모니터링", 2),
            ("컴플라이언스 위반건수", "건", "연간신규", "모니터링", 1),
            ("고객 민원 발생건수", "건", "연간신규", "모니터링", 1),
        ]
    },
    "영업추진2그룹": {
        "code": "SL2", "dept": "영업추진2부",
        "kpis": [
            ("중소기업 대출잔액", "억원", "총량", "KPI", 8),
            ("소호 대출 실행액", "억원", "연간신규", "KPI", 7),
            ("중소기업 신규 거래처", "社", "연간신규", "KPI", 7),
            ("소호 신규 거래처", "社", "연간신규", "KPI", 6),
            ("기업 수신 평잔", "억원", "총량", "KPI", 6),
            ("급여이체 기업 수", "社", "총량", "KPI", 5),
            ("기업 비이자수익", "억원", "연간신규", "KPI", 5),
            ("여신 건전성 비율", "%", "비율", "KPI", 5),
            ("신규 여신심사 처리시간", "일", "비율", "전략과제", 5),
            ("우량 여신 비율", "%", "비율", "전략과제", 5),
            ("법인카드 발급수", "좌", "연간신규", "전략과제", 4),
            ("기업 인터넷뱅킹 가입률", "%", "비율", "전략과제", 4),
            ("외환 거래 기업수", "社", "총량", "전략과제", 4),
            ("수출입 실적", "억원", "연간신규", "전략과제", 3),
            ("기업 모바일뱅킹 MAU", "명", "총량", "모니터링", 3),
            ("기업대출 연체율", "%", "비율", "모니터링", 3),
            ("소호 고객 이탈수", "社", "연간이탈", "모니터링", 3),
            ("B2B 플랫폼 거래건수", "건", "연간신규", "모니터링", 2),
            ("기업대출 NPL비율", "%", "비율", "모니터링", 2),
            ("영업점 기업고객 상담건수", "건", "연간신규", "모니터링", 3),
        ]
    },
    "기관·제휴영업그룹": {
        "code": "INS", "dept": "기관영업부",
        "kpis": [
            ("나라사랑카드 발급좌수", "좌", "연간신규", "KPI", 8),
            ("군간부 급여이체 고객수", "명", "총량", "KPI", 7),
            ("군인전용대출 취급액", "억원", "연간신규", "KPI", 7),
            ("대학생 체크카드 발급좌수", "좌", "연간신규", "KPI", 6),
            ("헤이영포유 이용고객", "명", "총량", "KPI", 6),
            ("헤이영포유 등록금 수납액", "억원", "연간신규", "KPI", 5),
            ("공공기관 수신잔액", "억원", "총량", "KPI", 5),
            ("제휴카드 발급건수", "건", "연간신규", "KPI", 5),
            ("기관 급여이체 건수", "건", "연간신규", "전략과제", 5),
            ("공공조달 수수료 수익", "억원", "연간신규", "전략과제", 5),
            ("학교 급식 결제 거래액", "억원", "연간신규", "전략과제", 4),
            ("의료기관 제휴 수", "社", "연간신규", "전략과제", 4),
            ("군인연금 수급고객", "명", "총량", "전략과제", 4),
            ("지자체 협약 건수", "건", "연간신규", "전략과제", 3),
            ("기관 고객 만족도", "점", "비율", "모니터링", 3),
            ("제휴 파트너 수", "社", "총량", "모니터링", 3),
            ("공공기관 신규 계약", "건", "연간신규", "모니터링", 3),
            ("기관대출 연체율", "%", "비율", "모니터링", 2),
            ("제휴 마케팅 ROI", "%", "비율", "모니터링", 2),
            ("기관 민원 건수", "건", "연간신규", "모니터링", 3),
        ]
    },
    "고객솔루션그룹": {
        "code": "CSG", "dept": "고객솔루션부",
        "kpis": [
            ("실거래고객수", "명", "총량", "KPI", 8),
            ("활동성고객수", "명", "총량", "KPI", 7),
            ("급여수급고객", "명", "총량", "KPI", 7),
            ("카드결제고객 (신한카드)", "명", "총량", "KPI", 6),
            ("모임통장 회원 고객수", "명", "총량", "KPI", 6),
            ("시니어 공적연금 수급고객", "명", "총량", "KPI", 5),
            ("외국인 거래고객수", "명", "총량", "KPI", 5),
            ("미성년자 거래고객수", "명", "총량", "KPI", 5),
            ("SOL 앱 MAU", "명", "총량", "전략과제", 5),
            ("1억 고객수 Shift", "명", "연간신규", "전략과제", 5),
            ("실거래고객 (중소법인)", "社", "총량", "전략과제", 4),
            ("실거래고객 (SOHO)", "社", "총량", "전략과제", 4),
            ("우량여신신규 고객수 (외감)", "社", "연간신규", "전략과제", 4),
            ("디지털 채널 가입율", "%", "비율", "전략과제", 3),
            ("50+ 걸어요 가입고객수", "명", "연간신규", "모니터링", 3),
            ("모임통장 입금액", "억원", "총량", "모니터링", 3),
            ("고객 이탈 방어율", "%", "비율", "모니터링", 3),
            ("CRM 캠페인 응답률", "%", "비율", "모니터링", 2),
            ("비대면 계좌개설 비율", "%", "비율", "모니터링", 2),
            ("고객 VOC 처리율", "%", "비율", "모니터링", 3),
        ]
    },
    "자산관리솔루션그룹": {
        "code": "WMS", "dept": "투자솔루션부",
        "kpis": [
            ("1억 고객수", "명", "총량", "KPI", 8),
            ("5억 고객수", "명", "총량", "KPI", 7),
            ("공모펀드 시장지위", "위", "총량", "KPI", 7),
            ("퇴직연금 시장지위", "위", "총량", "KPI", 6),
            ("일임형ISA 시장지위", "위", "총량", "KPI", 6),
            ("고객 종합수익률 (펀드)_BM比", "%p", "비율", "KPI", 5),
            ("고객 종합수익률 (신탁)_BM比", "%p", "비율", "KPI", 5),
            ("투자상품 판매경험 직원수", "명", "총량", "KPI", 5),
            ("퇴직연금 전문가 인증 직원수", "명", "총량", "전략과제", 5),
            ("유언대용신탁 계약건수", "좌", "연간신규", "전략과제", 5),
            ("유언대용신탁 계약금액", "억원", "연간신규", "전략과제", 4),
            ("특화신탁 잔액", "억원", "총량", "전략과제", 4),
            ("퇴직연금 적립금", "억원", "총량", "전략과제", 4),
            ("ISA 가입자수", "명", "연간신규", "전략과제", 3),
            ("펀드 순유입액", "억원", "연간신규", "모니터링", 3),
            ("ELS 만기상환율", "%", "비율", "모니터링", 3),
            ("신탁 신규 수탁액", "억원", "연간신규", "모니터링", 3),
            ("연금저축 이전 유치", "억원", "연간신규", "모니터링", 2),
            ("로보어드바이저 이용건수", "건", "연간신규", "모니터링", 2),
            ("투자상품 민원건수", "건", "연간신규", "모니터링", 3),
        ]
    },
    "CIB그룹": {
        "code": "CIB", "dept": "기업금융부",
        "kpis": [
            ("IB 수수료 수익", "억원", "연간신규", "KPI", 8),
            ("대기업 대출잔액", "억원", "총량", "KPI", 8),
            ("프로젝트파이낸싱 실행액", "억원", "연간신규", "KPI", 7),
            ("신디케이트론 주선액", "억원", "연간신규", "KPI", 6),
            ("인수금융 실행액", "억원", "연간신규", "KPI", 6),
            ("구조화금융 수익", "억원", "연간신규", "KPI", 5),
            ("대기업 수신 평잔", "억원", "총량", "KPI", 5),
            ("대기업 비이자수익", "억원", "연간신규", "KPI", 5),
            ("IPO 주관 건수", "건", "연간신규", "전략과제", 5),
            ("M&A 자문 수익", "억원", "연간신규", "전략과제", 5),
            ("회사채 인수 실적", "억원", "연간신규", "전략과제", 4),
            ("ESG 금융 실행액", "억원", "연간신규", "전략과제", 4),
            ("대기업 FX 거래규모", "억원", "연간신규", "전략과제", 3),
            ("대기업 파생상품 거래액", "억원", "연간신규", "전략과제", 3),
            ("대기업 신규 거래처", "社", "연간신규", "모니터링", 3),
            ("대기업 여신 건전성", "%", "비율", "모니터링", 3),
            ("PF 연체율", "%", "비율", "모니터링", 3),
            ("IB 딜 파이프라인", "건", "총량", "모니터링", 2),
            ("CIB 고객 만족도", "점", "비율", "모니터링", 2),
            ("대기업 민원건수", "건", "연간신규", "모니터링", 3),
        ]
    },
    "자본시장그룹": {
        "code": "CMK", "dept": "자본시장부",
        "kpis": [
            ("유가증권 운용수익", "억원", "연간신규", "KPI", 8),
            ("채권 트레이딩 수익", "억원", "연간신규", "KPI", 8),
            ("외환 딜링 수익", "억원", "연간신규", "KPI", 7),
            ("파생상품 운용수익", "억원", "연간신규", "KPI", 6),
            ("ALM 금리리스크 관리(VaR)", "억원", "총량", "KPI", 6),
            ("유동성커버리지비율(LCR)", "%", "비율", "KPI", 5),
            ("채권 포트폴리오 듀레이션", "년", "총량", "KPI", 5),
            ("단기자금 운용수익", "억원", "연간신규", "KPI", 5),
            ("해외채권 투자수익", "억원", "연간신규", "전략과제", 5),
            ("대체투자 수익률", "%", "비율", "전략과제", 5),
            ("환헤지 비용 절감액", "억원", "연간신규", "전략과제", 4),
            ("시장리스크 한도 소진율", "%", "비율", "전략과제", 4),
            ("국채 입찰 낙찰률", "%", "비율", "전략과제", 3),
            ("RP 운용 평잔", "억원", "총량", "전략과제", 3),
            ("채권 시장조성 실적", "억원", "연간신규", "모니터링", 3),
            ("FX Swap 포지션", "억원", "총량", "모니터링", 3),
            ("운용 손익 변동성", "%", "비율", "모니터링", 3),
            ("일일 P&L 최대손실", "억원", "총량", "모니터링", 2),
            ("시스템 트레이딩 비중", "%", "비율", "모니터링", 2),
            ("규제자본 적정성", "%", "비율", "모니터링", 3),
        ]
    },
    "글로벌사업그룹": {
        "code": "GBL", "dept": "글로벌사업부",
        "kpis": [
            ("해외점포 순이익", "억원", "연간신규", "KPI", 8),
            ("해외 대출잔액", "억원", "총량", "KPI", 8),
            ("해외 수신잔액", "억원", "총량", "KPI", 7),
            ("해외 비이자수익", "억원", "연간신규", "KPI", 6),
            ("SOL Global MAU", "명", "총량", "KPI", 6),
            ("해외송금 이용건수", "건", "연간신규", "KPI", 5),
            ("해외 신규 거래고객", "명", "연간신규", "KPI", 5),
            ("해외점포 ROE", "%", "비율", "KPI", 5),
            ("베트남 법인 순이익", "억원", "연간신규", "전략과제", 5),
            ("일본 SBJ은행 순이익", "억원", "연간신규", "전략과제", 5),
            ("글로벌 디지털뱅킹 가입자", "명", "연간신규", "전략과제", 4),
            ("해외 무역금융 실행액", "억원", "연간신규", "전략과제", 4),
            ("글로벌 IB 수익", "억원", "연간신규", "전략과제", 3),
            ("해외점포 비용효율성(CIR)", "%", "비율", "전략과제", 3),
            ("해외 여신 건전성", "%", "비율", "모니터링", 3),
            ("해외 환리스크 익스포저", "억원", "총량", "모니터링", 3),
            ("해외 컴플라이언스 이슈", "건", "연간신규", "모니터링", 3),
            ("신규 진출국 검토 건수", "건", "연간신규", "모니터링", 2),
            ("해외 인력 현지화율", "%", "비율", "모니터링", 2),
            ("해외 고객 민원건수", "건", "연간신규", "모니터링", 3),
        ]
    },
}

THIN = Side(style='thin', color='AAAAAA')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
HEADER_FONT = Font(name='맑은 고딕', size=10, bold=True, color='FFFFFF')
BODY_FONT = Font(name='맑은 고딕', size=10)
GROUP_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
KPI_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
STRATEGY_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
MONITOR_FILL = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
RED_FONT = Font(name='맑은 고딕', size=10, color='CC0000')
GREEN_FONT = Font(name='맑은 고딕', size=10, color='006100')
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
RED_FILL_LIGHT = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)


def style_header(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def style_cell(cell, align=CENTER, font=BODY_FONT, fill=None):
    cell.font = font
    cell.alignment = align
    cell.border = BORDER
    if fill:
        cell.fill = fill


def tool_fill(tool):
    if tool == 'KPI':
        return KPI_FILL
    elif tool == '전략과제':
        return STRATEGY_FILL
    return MONITOR_FILL


def gen_target(unit, basis):
    if unit in ('%', '%p'):
        return round(random.uniform(50, 99), 1)
    elif unit == '위':
        return random.randint(1, 5)
    elif unit == '점':
        return round(random.uniform(70, 95), 1)
    elif unit == '년':
        return round(random.uniform(2, 5), 1)
    elif unit == '일':
        return round(random.uniform(1, 10), 1)
    elif unit in ('억원', '백만원'):
        return random.randint(100, 5000)
    elif unit in ('명', '社', '좌', '건'):
        return random.randint(100, 50000)
    return random.randint(100, 10000)


def gen_monthly_targets(annual, unit):
    if unit in ('%', '%p', '위', '점', '년', '일'):
        return [annual] * 12
    monthly = []
    weights = [0.07, 0.07, 0.08, 0.08, 0.08, 0.09, 0.08, 0.08, 0.09, 0.09, 0.09, 0.10]
    for w in weights:
        monthly.append(round(annual * w, 1))
    return monthly


def gen_actual(target, unit):
    if target == 0:
        return 0
    ratio = random.uniform(0.7, 1.15)
    val = target * ratio
    if unit in ('%', '%p', '점', '년', '일'):
        return round(val, 1)
    return round(val)


wb = openpyxl.Workbook()

# ━━━━ Sheet 1: KPI 목표설정 ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws1 = wb.active
ws1.title = "1.KPI목표설정"
ws1.sheet_properties.tabColor = '2F5496'

headers1 = [
    'NO', '그룹', '지표코드', 'Category', '지표Seg', '지표명', '단위',
    '실적산출기준', '관리Tool', '평가비중(%)',
    '연간목표', '1월', '2월', '3월', '4월', '5월', '6월',
    '7월', '8월', '9월', '10월', '11월', '12월',
    'Data Ownership (부서)', '수집방식', '비고'
]
for c, h in enumerate(headers1, 1):
    ws1.cell(row=1, column=c, value=h)
style_header(ws1, 1, len(headers1))

ws1.freeze_panes = 'G2'

row = 2
no = 1
kpi_master = []

for gname, gdata in GROUPS.items():
    seg_counter = 1
    for idx, (kpi_name, unit, basis, tool, weight) in enumerate(gdata['kpis']):
        code = f"{gdata['code']}-{seg_counter:04d}-{no:04d}"
        annual = gen_target(unit, basis)
        monthly = gen_monthly_targets(annual, unit)
        collect = 'AUTO' if basis in ('총량', '비율') and unit not in ('점',) else '수기'

        vals = [no, gname, code, gdata['code'], f"Seg{seg_counter:03d}", kpi_name, unit,
                basis, tool, weight, annual] + monthly + [gdata['dept'], collect, '']

        for c, v in enumerate(vals, 1):
            cell = ws1.cell(row=row, column=c, value=v)
            fill = tool_fill(tool) if c == 9 else None
            style_cell(cell, align=CENTER if c != 6 else LEFT, fill=fill)

        kpi_master.append({
            'no': no, 'group': gname, 'code': code, 'category': gdata['code'],
            'seg': f"Seg{seg_counter:03d}", 'name': kpi_name, 'unit': unit,
            'basis': basis, 'tool': tool, 'weight': weight,
            'annual': annual, 'monthly': monthly, 'dept': gdata['dept'], 'collect': collect
        })

        no += 1
        row += 1
        if (idx + 1) % 5 == 0:
            seg_counter += 1

col_widths_1 = [5, 16, 16, 8, 10, 30, 6, 10, 8, 10, 12,
                10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10,
                18, 8, 12]
for i, w in enumerate(col_widths_1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers1))}{row - 1}"

# ━━━━ Sheet 2: 실적데이터 (DB형식) ━━━━━━━━━━━━━━━━━━━━━━━━
ws2 = wb.create_sheet("2.실적데이터")
ws2.sheet_properties.tabColor = '548235'

headers2 = [
    'ID', 'YYYY', 'MM', 'YYYYMM', '그룹', '지표코드', 'Category', '지표Seg',
    '지표명', '단위', '실적산출기준', '관리Tool', '평가비중(%)',
    '목표', '실적', '달성률(%)',
    '수집방식', 'Data Ownership (그룹)', 'Data Ownership (부서)',
    '입력자', '입력일시', '승인상태', '승인자', '승인일시', '비고'
]
for c, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=c, value=h)
style_header(ws2, 1, len(headers2))

ws2.freeze_panes = 'F2'

db_row = 2
rec_id = 1

for kpi in kpi_master:
    for m in range(1, 13):
        target = kpi['monthly'][m - 1]
        actual = gen_actual(target, kpi['unit'])
        achievement = round((actual / target) * 100, 1) if target != 0 else 0

        vals = [
            rec_id, 2026, m, f"2026{m:02d}",
            kpi['group'], kpi['code'], kpi['category'], kpi['seg'],
            kpi['name'], kpi['unit'], kpi['basis'], kpi['tool'], kpi['weight'],
            target, actual, achievement,
            kpi['collect'], kpi['group'], kpi['dept'],
            '', '', '미승인', '', '', ''
        ]

        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=db_row, column=c, value=v)
            style_cell(cell, align=CENTER if c != 9 else LEFT)
            if c == 16:
                if achievement >= 100:
                    cell.fill = GREEN_FILL
                elif achievement >= 80:
                    cell.fill = YELLOW_FILL
                else:
                    cell.fill = RED_FILL_LIGHT

        rec_id += 1
        db_row += 1

col_widths_2 = [8, 6, 4, 8, 16, 16, 8, 10, 30, 6, 10, 8, 10,
                12, 12, 10, 8, 16, 16, 10, 16, 8, 10, 16, 12]
for i, w in enumerate(col_widths_2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers2))}{db_row - 1}"

# ━━━━ Sheet 3: 실적현황 (대시보드) ━━━━━━━━━━━━━━━━━━━━━━━━
ws3 = wb.create_sheet("3.실적현황")
ws3.sheet_properties.tabColor = 'BF8F00'

TITLE_FONT = Font(name='맑은 고딕', size=14, bold=True, color='2F5496')
ws3.cell(row=1, column=1, value="■ 2026년 전행 KPI 그룹별 실적현황").font = TITLE_FONT
ws3.merge_cells('A1:R1')
ws3.cell(row=1, column=1).alignment = Alignment(horizontal='left', vertical='center')

dash_headers = [
    'NO', '그룹', '지표코드', '지표명', '단위', '관리Tool', '비중(%)',
    '연간목표',
    '1월실적', '1월달성률',
    '2월실적', '2월달성률',
    '3월실적', '3월달성률',
    'YTD실적', 'YTD목표', 'YTD달성률',
    '상태'
]
for c, h in enumerate(dash_headers, 1):
    ws3.cell(row=3, column=c, value=h)
style_header(ws3, 3, len(dash_headers))

ws3.freeze_panes = 'E4'

dash_row = 4
dash_no = 1
group_start_rows = {}

for gname, gdata in GROUPS.items():
    group_start_rows[gname] = dash_row
    group_kpis = [k for k in kpi_master if k['group'] == gname]

    for kpi in group_kpis:
        jan_target = kpi['monthly'][0]
        feb_target = kpi['monthly'][1]
        mar_target = kpi['monthly'][2]
        jan_actual = gen_actual(jan_target, kpi['unit'])
        feb_actual = gen_actual(feb_target, kpi['unit'])
        mar_actual = gen_actual(mar_target, kpi['unit'])

        jan_ach = round(jan_actual / jan_target * 100, 1) if jan_target else 0
        feb_ach = round(feb_actual / feb_target * 100, 1) if feb_target else 0
        mar_ach = round(mar_actual / mar_target * 100, 1) if mar_target else 0

        if kpi['unit'] in ('%', '%p', '위', '점', '년', '일'):
            ytd_actual = mar_actual
            ytd_target = mar_target
        else:
            ytd_actual = jan_actual + feb_actual + mar_actual
            ytd_target = jan_target + feb_target + mar_target

        ytd_ach = round(ytd_actual / ytd_target * 100, 1) if ytd_target else 0

        if ytd_ach >= 100:
            status = "●"
            s_fill = GREEN_FILL
        elif ytd_ach >= 80:
            status = "●"
            s_fill = YELLOW_FILL
        else:
            status = "●"
            s_fill = RED_FILL_LIGHT

        vals = [
            dash_no, gname, kpi['code'], kpi['name'], kpi['unit'],
            kpi['tool'], kpi['weight'], kpi['annual'],
            jan_actual, jan_ach, feb_actual, feb_ach, mar_actual, mar_ach,
            ytd_actual, ytd_target, ytd_ach, status
        ]
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(row=dash_row, column=c, value=v)
            fill = tool_fill(kpi['tool']) if c == 6 else None
            if c in (10, 12, 14, 17):
                if v >= 100:
                    fill = GREEN_FILL
                elif v >= 80:
                    fill = YELLOW_FILL
                else:
                    fill = RED_FILL_LIGHT
            if c == 18:
                fill = s_fill
            style_cell(cell, align=CENTER if c != 4 else LEFT, fill=fill)
            if c in (10, 12, 14, 17):
                cell.number_format = '0.0"%"'

        dash_no += 1
        dash_row += 1

    sep = ws3.cell(row=dash_row, column=1)
    dash_row += 1

col_widths_3 = [5, 16, 16, 30, 6, 8, 7, 12,
                12, 9, 12, 9, 12, 9,
                12, 12, 9, 5]
for i, w in enumerate(col_widths_3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.auto_filter.ref = f"A3:{get_column_letter(len(dash_headers))}{dash_row - 1}"

# ━━━━ Sheet 4: 그룹요약 ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws4 = wb.create_sheet("4.그룹요약")
ws4.sheet_properties.tabColor = 'C00000'

ws4.cell(row=1, column=1, value="■ 2026년 그룹별 KPI 종합 달성현황").font = TITLE_FONT
ws4.merge_cells('A1:H1')

sum_headers = ['NO', '그룹', 'KPI 지표수', '가중평균 달성률(%)', '100%↑', '80~99%', '80%↓', '종합 등급']
for c, h in enumerate(sum_headers, 1):
    ws4.cell(row=3, column=c, value=h)
style_header(ws4, 3, len(sum_headers))

srow = 4
sno = 1
for gname in GROUPS:
    group_kpis = [k for k in kpi_master if k['group'] == gname]
    weighted_sum = 0
    total_weight = 0
    over100 = 0
    mid = 0
    under80 = 0
    for kpi in group_kpis:
        t = kpi['annual']
        a = gen_actual(t, kpi['unit'])
        ach = (a / t * 100) if t else 0
        weighted_sum += ach * kpi['weight']
        total_weight += kpi['weight']
        if ach >= 100:
            over100 += 1
        elif ach >= 80:
            mid += 1
        else:
            under80 += 1

    wavg = round(weighted_sum / total_weight, 1) if total_weight else 0
    if wavg >= 100:
        grade = 'S'
        gfill = GREEN_FILL
    elif wavg >= 90:
        grade = 'A'
        gfill = GREEN_FILL
    elif wavg >= 80:
        grade = 'B'
        gfill = YELLOW_FILL
    elif wavg >= 70:
        grade = 'C'
        gfill = YELLOW_FILL
    else:
        grade = 'D'
        gfill = RED_FILL_LIGHT

    vals = [sno, gname, len(group_kpis), wavg, over100, mid, under80, grade]
    for c, v in enumerate(vals, 1):
        cell = ws4.cell(row=srow, column=c, value=v)
        fill = gfill if c == 8 else GROUP_FILL if c == 2 else None
        style_cell(cell, align=CENTER, fill=fill)
    sno += 1
    srow += 1

col_widths_4 = [5, 20, 12, 18, 8, 10, 8, 10]
for i, w in enumerate(col_widths_4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

# ━━━━ 저장 ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "KPI_성과관리_시스템.xlsx")
wb.save(out)
print(f"OK: {out}")
print(f"총 KPI: {len(kpi_master)}개 (8그룹 x 20개)")
print(f"실적 레코드: {rec_id - 1}건 (160 KPI x 12개월)")
